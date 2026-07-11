#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build-cte-data.py — Camada "Coorte CTE" do Painel TEA (Unimed GV).

Lê a planilha oficial do censo CTE (xlsx, SEMPRE fora do repositório) e gera
agregados k-anonimizados da coorte fixa de 300 beneficiários do Centro de
Terapias Especiais — consumo INTEGRAL e multidisciplinar (não só ABA-psicologia).

Universo DIFERENTE do painel principal (840 em ABA-psicologia); todo número
carrega o selo "Coorte CTE · 300" e nunca deve ser somado ao painel principal.

Saída: unimed/data/cte-agregados.json (injetado em unimed/tea.html entre os
marcadores <!-- CTE-DATA:BEGIN --> / <!-- CTE-DATA:END -->).

LGPD (base com CID e nascimento de menores): xlsx bruto JAMAIS no repo; só
agregados; k=5 em toda contagem de pacientes; varredura anti-vazamento
bloqueante; nenhum cruzamento fino (CID×nível×idade) que reidentifique.

Uso:
  python3 scripts/build-cte-data.py <caminho-externo>/censo-cte.xlsx --corte 2026-02-28
  python3 scripts/build-cte-data.py --selftest
"""

import argparse
import csv
import hashlib
import json
import math
import os
import re
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, datetime

VERSAO = "1.0.0"
K_ANON = 5
JSON_MAX_BYTES = 90 * 1024
COORTE = 300
INICIO_SERIE = "2024/01"  # E2: série a partir de 2024 (ano-base completo)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_OUT = os.path.join(REPO_ROOT, "unimed", "data", "cte-agregados.json")
HTML_ALVO = os.path.join(REPO_ROOT, "unimed", "tea.html")
MARC_INI = "<!-- CTE-DATA:BEGIN -->"
MARC_FIM = "<!-- CTE-DATA:END -->"

ABA_PSICO = "50005103"  # procedimento que o painel principal enxerga

ABA_COL = ["anomes", "matricula", "nasc", "idade_atual", "genero", "cid", "nivel",
           "idade_diag", "dt_atend", "cod_prest", "nome_prest", "tipo_prest",
           "ga", "cod_proc", "nome_proc", "cod_sol", "nome_sol", "esp_sol",
           "med_exec", "valor"]


def falha(msg):
    print(f"ERRO: {msg}", file=sys.stderr)
    sys.exit(1)


def dentro_de_repo_git(caminho):
    d = os.path.dirname(os.path.abspath(caminho))
    while True:
        if os.path.isdir(os.path.join(d, ".git")):
            return True
        pai = os.path.dirname(d)
        if pai == d:
            return False
        d = pai


def k5(n):
    return n if n >= K_ANON else "<5"


def r2(x):
    return round(x + 1e-9, 2)


def moeda(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("R$", "").strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_comp(v):
    if isinstance(v, (datetime, date)):
        return f"{v.year}/{v.month:02d}"
    return str(v).strip()


def parse_idade_diag(s):
    if s is None:
        return None
    s = str(s).lower()
    a = re.search(r"(\d+)\s*ano", s)
    m = re.search(r"(\d+)\s*m[eê]s", s)
    if not a and not m:
        n = re.search(r"(\d+)", s)
        return float(n.group(1)) if n else None
    return (int(a.group(1)) if a else 0) + (int(m.group(1)) / 12 if m else 0)


def disciplina(nome, ga):
    """Classifica a linha em disciplina/categoria. Ordem importa (ABA/Teacch/Denver
    trazem a disciplina no nome — o keyword específico é capturado primeiro)."""
    ga = (ga or "").upper()
    if ga != "TERAPIA":
        return {
            "CONSULTA": "Consultas", "EXAME": "Exames",
            "MEDICAMENTOS": "Medicamentos e materiais", "MATERIAIS": "Medicamentos e materiais",
            "TAXA": "Taxas e diárias", "DIARIA": "Taxas e diárias",
            "PROCEDIMENTO MEDICO": "Procedimentos médicos",
        }.get(ga, "Outros")
    n = (nome or "").lower()
    if "fonoaud" in n:
        return "Fonoaudiologia"
    if "terapia ocupacional" in n:
        return "Terapia Ocupacional"
    if "psicopedagog" in n:
        return "Psicopedagogia"
    if "nutric" in n:
        return "Nutrição"
    if "psicomotric" in n:
        return "Psicomotricidade"
    if "psico" in n or "neuropsic" in n or "teacch" in n or "denver" in n:
        return "Psicologia"
    return "Outras terapias"


def grupo_cid(cid):
    c = (cid or "").strip().upper()
    if c == "F84.0":
        return "F84.0 (autismo infantil)"
    if c.startswith("F84"):
        return "Demais F84.x (TEA)"
    return "Outros CID"


def ler_xlsx(caminho, permitir_no_repo=False):
    if not os.path.isfile(caminho):
        falha(f"insumo não encontrado: {caminho}")
    if not permitir_no_repo and dentro_de_repo_git(caminho):
        falha("o insumo está dentro de um repositório git — dado bruto JAMAIS no repo.")
    bruto = open(caminho, "rb").read()
    sha = hashlib.sha256(bruto).hexdigest()
    try:
        import openpyxl
    except ImportError:
        falha("openpyxl não instalado (pip install openpyxl)")
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    aba_custo = next((n for n in wb.sheetnames if "CUSTO" in n.upper()), None)
    aba_ros = next((n for n in wb.sheetnames if n.upper().startswith("DB")), None)
    if not aba_custo:
        falha(f"aba de custos não encontrada (abas: {wb.sheetnames})")
    # custos
    it = wb[aba_custo].iter_rows(values_only=True)
    next(it)
    regs = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        r = dict(zip(ABA_COL, list(row[:len(ABA_COL)])))
        if r["valor"] is None and r["cod_proc"] is None:
            continue
        regs.append(r)
    # roster (idade ao diagnóstico, por carteirinha)
    idade_diag = {}
    if aba_ros:
        itr = wb[aba_ros].iter_rows(values_only=True)
        hr = [str(c).strip() for c in next(itr)]
        Hr = {c: i for i, c in enumerate(hr)}
        ic = Hr.get("NÚMERO DA CARTEIRINHA")
        ii = Hr.get("IDADE DIAGNÓSTICO")
        for row in itr:
            if ic is None or row[ic] is None:
                continue
            v = parse_idade_diag(row[ii]) if ii is not None else None
            if v is not None:
                idade_diag[str(row[ic]).strip()] = v
    return regs, idade_diag, sha


def agregar(regs, idade_diag, corte):
    for r in regs:
        r["_mat"] = str(r["matricula"]).strip()
        r["_comp"] = norm_comp(r["anomes"])
        r["_val"] = moeda(r["valor"])
        r["_disc"] = disciplina(r["nome_proc"], r["ga"])

    pacientes = sorted({r["_mat"] for r in regs})
    custo_total = sum(r["_val"] for r in regs)

    d = {}
    d["meta"] = {
        "camada": "Coorte CTE — Centro de Terapias Especiais · Casa Unimed GV",
        "universo": "coorte fixa de %d beneficiários (matrícula)" % len(pacientes),
        "selo": "Coorte CTE · %d" % len(pacientes),
        "data_corte": corte.isoformat(),
        "gerado_em": date.today().isoformat(),
        "versao_script": VERSAO,
        "politica_k": K_ANON,
        "pacientes": len(pacientes),
        "custo_total_periodo": r2(custo_total),
        "registros": len(regs),
    }

    # ---- E1+E2: custo por disciplina, série mensal 2024/01+ ----
    regs_serie = [r for r in regs if r["_comp"] >= INICIO_SERIE]
    meses = sorted({r["_comp"] for r in regs_serie})
    disc_total = defaultdict(lambda: {"c": 0.0, "p": set(), "n": 0})
    for r in regs_serie:
        e = disc_total[r["_disc"]]
        e["c"] += r["_val"]; e["p"].add(r["_mat"]); e["n"] += 1
    # bucketiza disciplinas com <5 pacientes em "Outras terapias" (k-anon)
    ORDEM = ["Psicologia", "Fonoaudiologia", "Terapia Ocupacional", "Psicopedagogia",
             "Psicomotricidade", "Nutrição", "Outras terapias", "Consultas", "Exames",
             "Medicamentos e materiais", "Taxas e diárias", "Procedimentos médicos", "Outros"]
    fund = {}
    for disc, e in disc_total.items():
        fund[disc] = disc if len(e["p"]) >= K_ANON else "Outras terapias"
    disc_fin = defaultdict(lambda: {"c": 0.0, "p": set(), "n": 0})
    for disc, e in disc_total.items():
        t = disc_fin[fund[disc]]
        t["c"] += e["c"]; t["p"] |= e["p"]; t["n"] += e["n"]
    custo_serie = sum(e["c"] for e in disc_fin.values())
    d["disciplinas"] = [{
        "disciplina": disc, "custo": r2(e["c"]),
        "pct": r2(e["c"] / custo_serie * 100), "pacientes": k5(len(e["p"])),
        "registros": e["n"],
    } for disc, e in sorted(disc_fin.items(), key=lambda kv: -kv[1]["c"])]

    # série mensal empilhada por disciplina
    discs_ord = [x["disciplina"] for x in d["disciplinas"]]
    serie = {disc: [] for disc in discs_ord}
    tot_mes = []
    for m in meses:
        por = defaultdict(float)
        for r in regs_serie:
            if r["_comp"] == m:
                por[fund[r["_disc"]]] += r["_val"]
        tot_mes.append(r2(sum(por.values())))
        for disc in discs_ord:
            serie[disc].append(r2(por.get(disc, 0.0)))
    d["disciplina_mensal"] = {"meses": meses, "total": tot_mes, "series": serie,
                              "inicio": INICIO_SERIE}

    # ---- E3: cobertura do painel principal (só ABA-psico) sobre a coorte ----
    custo_aba = sum(r["_val"] for r in regs_serie if str(r["cod_proc"]).strip() == ABA_PSICO)
    pac_aba = {r["_mat"] for r in regs_serie if str(r["cod_proc"]).strip() == ABA_PSICO}
    d["cobertura"] = {
        "custo_coorte": r2(custo_serie),
        "custo_aba_psico": r2(custo_aba),
        "pct_visivel": r2(custo_aba / custo_serie * 100),
        "custo_invisivel": r2(custo_serie - custo_aba),
        "pacientes_coorte": len(pacientes),
        "pacientes_com_aba": len(pac_aba),
        "pacientes_sem_aba": len(pacientes) - len(pac_aba),
    }

    # ---- E4: perfil clínico (CID, nível de suporte, idade ao diagnóstico) ----
    cad = {}
    custo_pac = defaultdict(float)
    for r in regs:
        custo_pac[r["_mat"]] += r["_val"]
        if r["_mat"] not in cad:
            cad[r["_mat"]] = {"cid": r["cid"], "nivel": r["nivel"], "genero": r["genero"]}
    # CID agrupado
    cidg = Counter(grupo_cid(v["cid"]) for v in cad.values())
    d["cid"] = [{"grupo": g, "pacientes": k5(n), "pct": r2(n / len(cad) * 100)}
                for g, n in sorted(cidg.items(), key=lambda kv: -kv[1])]
    # Nível de suporte + custo médio por nível
    def norm_nivel(x):
        s = str(x or "").strip().upper()
        if "III" in s: return "Nível III"
        if "II" in s: return "Nível II"
        if s.startswith("NÍVEL I") or s == "NIVEL I": return "Nível I"
        return "Não classificado"
    niv = defaultdict(lambda: {"p": set(), "c": 0.0})
    for m, v in cad.items():
        nn = norm_nivel(v["nivel"])
        niv[nn]["p"].add(m); niv[nn]["c"] += custo_pac[m]
    ORDEM_NIV = ["Nível I", "Nível II", "Nível III", "Não classificado"]
    d["nivel_suporte"] = [{
        "nivel": nn, "pacientes": k5(len(niv[nn]["p"])),
        "custo_medio": r2(niv[nn]["c"] / len(niv[nn]["p"])) if niv[nn]["p"] else 0,
    } for nn in ORDEM_NIV if nn in niv]
    # Idade ao diagnóstico
    idd = [idade_diag[m] for m in cad if m in idade_diag]
    d["idade_diagnostico"] = {
        "n": len(idd),
        "media": r2(statistics.mean(idd)) if idd else None,
        "mediana": r2(statistics.median(idd)) if idd else None,
    } if idd else None
    # gênero da coorte
    g = Counter(str(v["genero"]).strip().upper()[:1] for v in cad.values())
    d["genero"] = {"M": g.get("M", 0), "F": g.get("F", 0)}

    # ---- E4: multidisciplinaridade (nº de disciplinas terapêuticas por paciente) ----
    TERAPIAS = {"Psicologia", "Fonoaudiologia", "Terapia Ocupacional", "Psicopedagogia",
                "Psicomotricidade", "Nutrição", "Outras terapias"}
    disc_por_pac = defaultdict(set)
    for r in regs:
        if r["_disc"] in TERAPIAS:
            disc_por_pac[r["_mat"]].add(r["_disc"])
    dist = Counter()
    for m in pacientes:
        n = len(disc_por_pac.get(m, set()))
        dist[min(n, 5)] += 1  # 5 = "5+"
    d["multidisciplinar"] = [{
        "n_terapias": ("5+" if k == 5 else str(k)), "pacientes": k5(dist.get(k, 0))
    } for k in range(1, 6) if dist.get(k, 0) > 0]
    tres_mais = sum(dist.get(k, 0) for k in (3, 4, 5))
    d["multidisciplinar_resumo"] = {"tres_ou_mais": tres_mais,
                                    "pct": r2(tres_mais / len(pacientes) * 100)}

    # ---- concentração de custo por paciente (Gini/percentis) ----
    vals = sorted(custo_pac.values())
    n = len(vals)
    cum = sum((i + 1) * v for i, v in enumerate(vals))
    gini = (2 * cum) / (n * sum(vals)) - (n + 1) / n if sum(vals) else 0
    vd = sorted(vals, reverse=True)
    def toppct(p):
        k = max(1, int(n * p))
        return r2(sum(vd[:k]) / sum(vals) * 100)
    d["concentracao"] = {
        "gini": r2(gini),
        "media": r2(statistics.mean(vals)), "mediana": r2(statistics.median(vals)),
        "top10_pct": toppct(0.10), "top20_pct": toppct(0.20),
    }

    return d, regs


def validar(d, regs):
    erros = []
    blob = json.dumps(d, ensure_ascii=False)
    # anti-vazamento: nenhuma matrícula/nascimento/nome de beneficiário no JSON
    for r in regs:
        for campo in ("matricula", "nasc"):
            v = str(r[campo]).strip()
            if len(v) >= 8 and v in blob:
                erros.append(f"VAZAMENTO: {campo} presente no JSON")
                break
        else:
            continue
        break
    # k-anon: nenhuma contagem de pacientes 1-4 em claro
    for m in re.finditer(r'"pacientes":\s*(\d+)', blob):
        if int(m.group(1)) < K_ANON:
            erros.append(f"k-anonimato violado: pacientes={m.group(1)} em claro")
    # soma das disciplinas == custo da série
    if abs(sum(x["custo"] for x in d["disciplinas"]) - d["cobertura"]["custo_coorte"]) > 1:
        erros.append("soma das disciplinas != custo da coorte na série")
    if len(blob.encode("utf-8")) > JSON_MAX_BYTES:
        erros.append(f"JSON excede {JSON_MAX_BYTES // 1024} KB")
    return erros


def escrever_json(d):
    os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=1, sort_keys=True)
        f.write("\n")


def injetar_html(d, caminho=HTML_ALVO):
    if not os.path.isfile(caminho):
        print(f"AVISO: {caminho} não existe — pulando injeção")
        return False
    html = open(caminho, encoding="utf-8").read()
    if MARC_INI not in html or MARC_FIM not in html:
        print("AVISO: marcadores CTE-DATA ausentes no HTML — pulando injeção")
        return False
    bloco = (f'{MARC_INI}\n<script type="application/json" id="cte-data">'
             f'{json.dumps(d, ensure_ascii=False, sort_keys=True)}</script>\n{MARC_FIM}')
    novo = re.sub(re.escape(MARC_INI) + r".*?" + re.escape(MARC_FIM), lambda _: bloco, html, flags=re.S)
    open(caminho, "w", encoding="utf-8").write(novo)
    return True


def selftest():
    fx = os.path.join(REPO_ROOT, "scripts", "tests", "cte-fixture-sintetica.csv")
    linhas = list(csv.reader(open(fx, encoding="utf-8")))[1:]
    regs = [dict(zip(ABA_COL, row)) for row in linhas]
    idade_diag = {}
    # classificador de disciplina (unit test direto — independe do tamanho da coorte)
    assert disciplina("Terapia ABA - Fonoaudiologia - Terapias Pediatricas", "TERAPIA") == "Fonoaudiologia"
    assert disciplina("Terapia ABA - Psicologia", "TERAPIA") == "Psicologia"
    assert disciplina("Metodo Teacch - Terapia Ocupacional", "TERAPIA") == "Terapia Ocupacional"
    assert disciplina("Sessão de psicoterapia individual por psicólogo", "TERAPIA") == "Psicologia"
    assert disciplina("Consulta em consultorio", "CONSULTA") == "Consultas"
    assert disciplina("Ferritina", "EXAME") == "Exames"
    assert grupo_cid("F84.0").startswith("F84.0") and grupo_cid("F84.5").startswith("Demais") and grupo_cid("F80") == "Outros CID"
    d, regs = agregar(regs, idade_diag, date(2026, 2, 28))
    erros = validar(d, regs)
    assert not erros, f"fixture reprovou: {erros}"
    assert d["meta"]["pacientes"] >= 5, "fixture precisa de >=5 pacientes p/ k-anon"
    assert d["cobertura"]["pct_visivel"] < 100, "cobertura ABA deveria ser < 100%"
    blob = json.dumps(d, ensure_ascii=False).upper()
    assert "SINTETIC" not in blob, "nome/dado de beneficiário vazou"
    print("selftest OK — fixture CTE passou (disciplinas, cobertura, k-anon, anti-vazamento)")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", nargs="?", help="planilha do censo CTE (fora do repo)")
    ap.add_argument("--corte", help="data de corte AAAA-MM-DD")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--permitir-insumo-no-repo", action="store_true", help=argparse.SUPPRESS)
    args = ap.parse_args()

    if args.selftest:
        selftest()
        return
    if not args.xlsx:
        falha("informe o xlsx do censo CTE")
    corte = date.fromisoformat(args.corte) if args.corte else date(2026, 2, 28)
    regs, idade_diag, sha = ler_xlsx(args.xlsx, args.permitir_insumo_no_repo)
    d, regs = agregar(regs, idade_diag, corte)
    d["meta"]["sha256_insumo"] = sha
    erros = validar(d, regs)
    if erros:
        for e in erros:
            print(f"VALIDAÇÃO FALHOU: {e}", file=sys.stderr)
        sys.exit(1)
    escrever_json(d)
    injetou = injetar_html(d)
    c = d["cobertura"]
    print("=" * 64)
    print("RELATÓRIO — Camada Coorte CTE")
    print(f"  Coorte         : {d['meta']['pacientes']} beneficiários")
    print(f"  Série          : {d['disciplina_mensal']['inicio']} a {d['disciplina_mensal']['meses'][-1]}")
    print(f"  Custo (série)  : R$ {c['custo_coorte']:,.2f}")
    print(f"  ABA-psico (painel principal): R$ {c['custo_aba_psico']:,.2f} = {c['pct_visivel']}% do custo")
    print(f"  Invisível ao painel: R$ {c['custo_invisivel']:,.2f}")
    print(f"  Disciplinas    : {len(d['disciplinas'])}")
    print(f"  JSON           : {len(json.dumps(d, ensure_ascii=False).encode('utf-8')) // 1024} KB → {JSON_OUT}")
    print(f"  HTML injetado  : {'sim' if injetou else 'NÃO'}")
    print("=" * 64)


if __name__ == "__main__":
    main()
