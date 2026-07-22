#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline auditável do Painel Terapias Especiais — Unimed Governador Valadares.

Entrada de produção (artefato privado, fora de qualquer Git):
  python scripts/build-tea-data.py <diretorio-com-19-fontes-xlsx>
    --matriz <Matriz_Analitica_Terapias_Especiais_2025.xlsx>
    --saida <diretorio-privado>/tea-2025-2026.json
    --corte 2026-06-30

O diretório, a Matriz e a saída devem estar fora de qualquer repositório Git.
O pipeline emite agregados completos para a camada privada autenticada, mantém
os 19 hashes de origem e reconcilia 2025 com a Matriz. Nome, matrícula, data de
nascimento e guia de beneficiário nunca integram o JSON.

Comandos auxiliares:
  python scripts/build-tea-data.py --selftest
  python scripts/build-tea-data.py --emit-p
"""

import argparse
import hashlib
import json
import math
import os
import re
import statistics
import sys
import tempfile
import unicodedata
from collections import Counter, defaultdict, deque
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path


VERSAO = "5.0.0"

# Contrato v4 (PRD escopos comparáveis): parâmetros das referências por
# disciplina — mesmos limiares publicados em Método e Fontes.
MIN_SESSOES_REFERENCIA = Decimal(100)
MIN_PRESTADORES_REFERENCIA = 5
JSON_MAX_BYTES = 8 * 1024 * 1024

REPO_ROOT = Path(__file__).resolve().parents[1]
HTML_ALVO = REPO_ROOT / "unimed" / "tea.html"
P_DIR = REPO_ROOT / "p" / "painel-tea"

CORTE_PADRAO = date(2026, 6, 30)
INICIO_CANONICO = "2025/01"
FIM_CANONICO = "2026/06"


def sequencia_meses(inicio, fim):
    yi, mi = map(int, inicio.split("/"))
    yf, mf = map(int, fim.split("/"))
    meses = []
    y, m = yi, mi
    while (y, m) <= (yf, mf):
        meses.append(f"{y:04d}/{m:02d}")
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return tuple(meses)


MESES_CANONICOS = sequencia_meses(INICIO_CANONICO, FIM_CANONICO)
MESES_CANONICOS_SET = set(MESES_CANONICOS)
MESES_PECS_ESPERADOS = {
    "2025/01", "2025/02", "2025/03", "2025/05", "2025/08",
    "2025/09", "2025/11", "2025/12", "2026/01", "2026/02",
    "2026/03", "2026/05", "2026/06",
}

# O cabeçalho abaixo reproduz literalmente as 19 fontes oficiais recebidas.
# A grafia sem acento de "Matricula" pertence ao schema de origem.
CABECALHO_FONTE = [
    "Ano/Mês",
    "Matricula Beneficiário",
    "Nome Beneficiário",
    "Data Nascimento",
    "Gênero Beneficiário",
    "Numero Guia Principal",
    "Data Atendimento",
    "Código Solicitante",
    "Nome Solicitante",
    "CBOS Solicitante",
    "Tipo Guia",
    "Código Procedimento",
    "Nome Procedimento",
    "Tipo Procedimento GA",
    "Código Prestador",
    "Nome Prestador",
    "Tipo Prestador",
    "Quantidade Procedimento",
    "Pagamento/Despesa",
]

CHAVES_FONTE = [
    "anomes", "matricula", "nome", "nasc", "genero", "guia",
    "dt_atend", "cod_sol", "nome_sol", "cbos", "tipo_guia",
    "cod_proc", "nome_proc", "tipo_proc", "cod_prest",
    "nome_prest", "tipo_prest", "qtd", "valor",
]

# 19 códigos de origem -> 18 terapias clínicas. O código de AT integra ABA
# Psicologia e permanece identificável nos agregados de composição.
MAPA_CODIGOS = {
    "50005103": {
        "id": "aba-psicologia",
        "terapia": "Terapia ABA — Psicologia",
        "especialidade": "Psicologia",
        "metodo": "ABA",
    },
    "50000519": {
        "id": "aba-psicologia",
        "terapia": "Terapia ABA — Psicologia",
        "especialidade": "Psicologia",
        "metodo": "ABA",
        "at": True,
    },
    "50005189": {
        "id": "aba-fonoaudiologia",
        "terapia": "Terapia ABA — Fonoaudiologia",
        "especialidade": "Fonoaudiologia",
        "metodo": "ABA",
    },
    "50005170": {
        "id": "aba-terapia-ocupacional",
        "terapia": "Terapia ABA — Terapia Ocupacional",
        "especialidade": "Terapia Ocupacional",
        "metodo": "ABA",
    },
    "50005278": {
        "id": "psicopedagogia",
        "terapia": "Psicopedagogia",
        "especialidade": "Psicopedagogia",
        "metodo": "Psicopedagogia",
    },
    "50005375": {
        "id": "ans-centros-referencia",
        "terapia": "Terapias especiais (parecer ANS, centros de referência)",
        "especialidade": "Multidisciplinar",
        "metodo": "ANS — centros de referência",
    },
    "50005235": {
        "id": "denver-terapia-ocupacional",
        "terapia": "Método Denver — Terapia Ocupacional",
        "especialidade": "Terapia Ocupacional",
        "metodo": "Denver",
    },
    "50005227": {
        "id": "denver-psicologia",
        "terapia": "Método Denver — Psicologia",
        "especialidade": "Psicologia",
        "metodo": "Denver",
    },
    "50005243": {
        "id": "denver-fonoaudiologia",
        "terapia": "Método Denver — Fonoaudiologia",
        "especialidade": "Fonoaudiologia",
        "metodo": "Denver",
    },
    "50005138": {
        "id": "teacch-psicologia",
        "terapia": "Método TEACCH — Psicologia",
        "especialidade": "Psicologia",
        "metodo": "TEACCH",
    },
    "50005219": {
        "id": "teacch-fonoaudiologia",
        "terapia": "Método TEACCH — Fonoaudiologia",
        "especialidade": "Fonoaudiologia",
        "metodo": "TEACCH",
    },
    "50005200": {
        "id": "teacch-terapia-ocupacional",
        "terapia": "Método TEACCH — Terapia Ocupacional",
        "especialidade": "Terapia Ocupacional",
        "metodo": "TEACCH",
    },
    "50005197": {
        "id": "bobath-terapia-ocupacional",
        "terapia": "Método Bobath — Terapia Ocupacional Neurológica",
        "especialidade": "Terapia Ocupacional",
        "metodo": "Bobath",
    },
    "50005251": {
        "id": "bobath-fonoaudiologia",
        "terapia": "Método Bobath — Fonoaudiologia",
        "especialidade": "Fonoaudiologia",
        "metodo": "Bobath",
    },
    "50005260": {
        "id": "integracao-sensorial",
        "terapia": "Integração Sensorial",
        "especialidade": "Terapia Ocupacional",
        "metodo": "Integração Sensorial",
    },
    "50005340": {
        "id": "tgd-terapia-ocupacional",
        "terapia": "Terapeuta Ocupacional — TGD",
        "especialidade": "Terapia Ocupacional",
        "metodo": "TGD",
    },
    "50005286": {
        "id": "tgd-psicologia",
        "terapia": "Psicólogo — TGD",
        "especialidade": "Psicologia",
        "metodo": "TGD",
    },
    "50005308": {
        "id": "tgd-fonoaudiologia",
        "terapia": "Fonoaudiólogo — TGD",
        "especialidade": "Fonoaudiologia",
        "metodo": "TGD",
    },
    "50005146": {
        "id": "pecs-fonoaudiologia",
        "terapia": "Método PECS — Fonoaudiologia",
        "especialidade": "Fonoaudiologia",
        "metodo": "PECS",
    },
}

COD_AT = "50000519"
ID_PECS = "pecs-fonoaudiologia"
PEDIAKIDS = "PEDIAKIDS CURSOS E SERVICOS MEDICOS LTDA"

MERGE_CNPJ = {
    "ANA CAROLINA OLIVEIRA FARIA FALCAO LTDA",
    "CRIAR E CRESCER TEEN LTDA",
}
NOME_MERGE_CNPJ = "CRIAR E CRESCER TEEN LTDA"

MAPA_CANAL = {
    "ATENDIMENTO ESPECIALIZADO": "Casa Unimed",
    "ATENDIMENTOS FORNECEDORES": "Negociação Direta",
    "CLINICA ESPECIALIZADA": "Rede Credenciada",
    "OUTRAS ESPEC. EXCETO MEDICO": "Negociação Direta (P.F.)",
    "REEMBOLSO": "Reembolso",
    "UNIMED NACIONAL": "Intercâmbio",
    "UNIMED REGIONAL": "Intercâmbio",
}

CREDENCIADA_OVERRIDE = {
    "RESINTO ESPACO INTEGRADO LTDA",
    "CLINICA PSICONEURO GV LTDA",
    "ANIMA CENTRO DE TERAPIAS MULTIDISCIPLINAR LTDA",
    "CUIDARIM ESPACO DE DESENVOLVIMENTO INFANTIL LTDA",
    "CASA AURORA TERAPIAS LTDA",
}

# Contrato v5: todo registro com Tipo Prestador = ATENDIMENTO ESPECIALIZADO é
# atendimento da Casa Unimed e consolida na entidade única publicada. A regra
# vale por REGISTRO: o mesmo profissional permanece prestador individual nos
# registros dos demais canais.
CASA_UNIMED = "CASA UNIMED"
TIPO_CASA_UNIMED = "ATENDIMENTO ESPECIALIZADO"
CANAL_CASA_UNIMED = "Casa Unimed"

# Contrato v5: kits terapêuticos — mesma taxonomia do painel. Cada criança é
# classificada pelo CONJUNTO de disciplinas utilizadas no recorte; os kits são
# mutuamente exclusivos e exaustivos por construção. Célula com menos de
# K_MINIMO_KIT crianças não publica (k-anonimato).
K_MINIMO_KIT = 5
DISC_PSI = "Psicologia"
DISC_FONO = "Fonoaudiologia"
DISC_TO = "Terapia Ocupacional"
TRIPE_DISCIPLINAS = frozenset({DISC_PSI, DISC_FONO, DISC_TO})
KITS = (
    {
        "id": "tripe",
        "rotulo": "Psicologia + Fonoaudiologia + Terapia Ocupacional",
        "criterio": "exatamente as três disciplinas principais em conjunto",
    },
    {
        "id": "tripe-ampliado",
        "rotulo": "Psicologia + Fonoaudiologia + Terapia Ocupacional + outras",
        "criterio": "as três disciplinas principais mais uma ou mais complementares",
    },
    {
        "id": "duas",
        "rotulo": "Duas Terapias",
        "criterio": "exatamente duas disciplinas, em qualquer combinação",
    },
    {
        "id": "multiplas-sem-tripe",
        "rotulo": "Três ou Mais — Outras Combinações",
        "criterio": "três ou mais disciplinas sem conter as três principais",
    },
    {
        "id": "unica",
        "rotulo": "Terapia Única",
        "criterio": "uma única disciplina no recorte",
    },
)

FAIXAS_ETARIAS = ("0–2", "3–5", "6–8", "9–11", "12–14", "15–17", "18+")
BUCKETS_INTENSIDADE = ("<3", "3–4", "5–8", "9+")
CENTAVO = Decimal("0.01")

class ErroPipeline(RuntimeError):
    pass


def d_nome(valor):
    """Rótulo de prestador/solicitante: CAIXA ALTA sem acento."""
    texto = "" if valor is None else str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).upper()


def chave_rotulo(valor):
    return re.sub(r"[^A-Z0-9]+", " ", d_nome(valor)).strip()


def texto_celula(valor):
    if valor is None:
        return ""
    if isinstance(valor, (date, datetime)):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def codigo_procedimento(valor):
    if isinstance(valor, (int, float, Decimal)) and not isinstance(valor, bool):
        d = Decimal(str(valor))
        if d == d.to_integral_value():
            return str(int(d))
    texto = texto_celula(valor)
    if re.fullmatch(r"\d{8}(?:\.0+)?", texto):
        return texto.split(".")[0]
    raise ErroPipeline(f"código de procedimento inválido: {texto!r}")


def normalizar_anomes(valor):
    if isinstance(valor, (date, datetime)):
        return f"{valor.year:04d}/{valor.month:02d}"
    texto = texto_celula(valor)
    m = re.fullmatch(r"(\d{4})[/-](\d{1,2})(?:[/-]\d{1,2})?", texto)
    if m:
        ano, mes = int(m.group(1)), int(m.group(2))
    else:
        m = re.fullmatch(r"(\d{1,2})[/-](\d{4})", texto)
        if not m:
            raise ErroPipeline(f"competência inválida: {texto!r}")
        mes, ano = int(m.group(1)), int(m.group(2))
    if not 1 <= mes <= 12:
        raise ErroPipeline(f"competência inválida: {texto!r}")
    return f"{ano:04d}/{mes:02d}"


def normalizar_data(valor, campo):
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    texto = texto_celula(valor)
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto[:10], formato).date().isoformat()
        except ValueError:
            pass
    raise ErroPipeline(f"{campo} inválida: {texto!r}")


def decimal_celula(valor, campo):
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return Decimal(str(valor))
    texto = texto_celula(valor).replace("\u00a0", "").replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ErroPipeline(f"{campo} inválido: {valor!r}") from exc


def dinheiro(valor):
    if valor is None:
        return None
    return float(Decimal(valor).quantize(CENTAVO, rounding=ROUND_HALF_UP))


def numero(valor):
    if valor is None:
        return None
    d = Decimal(valor)
    return int(d) if d == d.to_integral_value() else float(d)


def razao(numerador, denominador):
    if not denominador:
        return None
    return dinheiro(Decimal(numerador) / Decimal(denominador))


def sha256_arquivo(caminho):
    h = hashlib.sha256()
    with open(caminho, "rb") as f:
        for bloco in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()


def ancestral_git(caminho):
    p = Path(caminho).resolve(strict=True)
    if p.is_file():
        p = p.parent
    while True:
        if (p / ".git").exists():
            return p
        if p.parent == p:
            return None
        p = p.parent


def exigir_fora_de_git(caminho, rotulo):
    real = Path(caminho).resolve(strict=True)
    git = ancestral_git(real)
    if git is not None:
        raise ErroPipeline(
            f"{rotulo} está dentro do repositório Git {git}; "
            "dados brutos devem permanecer fora de qualquer Git"
        )
    return real


def exigir_saida_fora_de_git(caminho):
    """Valida um destino ainda inexistente sem criar nada no repositório."""
    destino = Path(caminho).expanduser().resolve(strict=False)
    if destino.exists() and destino.is_dir():
        raise ErroPipeline("--saida deve apontar para um arquivo JSON")
    if destino.suffix.lower() != ".json":
        raise ErroPipeline("--saida deve usar a extensão .json")
    pai = destino.parent
    if not pai.is_dir():
        raise ErroPipeline(
            f"diretório de --saida não existe: {pai}; crie-o fora de qualquer Git"
        )
    git = ancestral_git(pai)
    if git is not None:
        raise ErroPipeline(
            f"--saida está dentro do repositório Git {git}; "
            "o artefato completo deve permanecer na camada privada"
        )
    return destino


def catalogo_terapias():
    por_id = {}
    for codigo, regra in MAPA_CODIGOS.items():
        item = por_id.setdefault(
            regra["id"],
            {
                "id": regra["id"],
                "terapia": regra["terapia"],
                "especialidade": regra["especialidade"],
                "metodo": regra["metodo"],
                "codigos": [],
            },
        )
        item["codigos"].append(codigo)
    for item in por_id.values():
        item["codigos"].sort(key=lambda c: (c == COD_AT, c))
    return por_id


CATALOGO_TERAPIAS = catalogo_terapias()


def prestador_consolidado(nome_fonte):
    return NOME_MERGE_CNPJ if nome_fonte in MERGE_CNPJ else nome_fonte


def canal_do(tipo_prestador, nome_fonte):
    if nome_fonte in CREDENCIADA_OVERRIDE:
        return "Rede Credenciada"
    tipo = d_nome(tipo_prestador)
    if tipo not in MAPA_CANAL:
        raise ErroPipeline(
            f"Tipo Prestador desconhecido: {tipo_prestador!r}; atualize o mapa auditável"
        )
    return MAPA_CANAL[tipo]


def chave_paciente(registro):
    # Regra oficial: valor de Nome + data de nascimento; sem matrícula e sem
    # normalização fonética, retirada de acento ou aproximação de grafia.
    return registro["nome"], registro["nasc"]


def localizar_aba_dados(wb, arquivo):
    candidatas = []
    for ws in wb.worksheets:
        for numero_linha, row in enumerate(
            ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1
        ):
            valores = [texto_celula(v) for v in row]
            while valores and valores[-1] == "":
                valores.pop()
            if valores and valores[0] == "Ano/Mês":
                if valores != CABECALHO_FONTE:
                    raise ErroPipeline(
                        f"{arquivo.name}: schema divergente na aba {ws.title!r}, "
                        f"linha {numero_linha}; esperado exatamente 19 colunas"
                    )
                candidatas.append((ws, numero_linha))
                break
    if len(candidatas) != 1:
        raise ErroPipeline(
            f"{arquivo.name}: esperado um único cabeçalho 'Ano/Mês'; "
            f"encontrados {len(candidatas)}"
        )
    return candidatas[0]


def normalizar_registro(valores, arquivo, numero_linha):
    bruto = dict(zip(CHAVES_FONTE, valores))
    try:
        anomes = normalizar_anomes(bruto["anomes"])
        nasc = normalizar_data(bruto["nasc"], "Data Nascimento")
        codigo = codigo_procedimento(bruto["cod_proc"])
        qtd = decimal_celula(bruto["qtd"], "Quantidade Procedimento")
        valor = decimal_celula(bruto["valor"], "Pagamento/Despesa")
    except ErroPipeline as exc:
        raise ErroPipeline(f"{arquivo.name}, linha {numero_linha}: {exc}") from exc

    if codigo not in MAPA_CODIGOS:
        raise ErroPipeline(
            f"{arquivo.name}, linha {numero_linha}: código fora do catálogo: {codigo}"
        )
    if qtd < 0:
        raise ErroPipeline(
            f"{arquivo.name}, linha {numero_linha}: Quantidade Procedimento negativa"
        )

    nome = texto_celula(bruto["nome"])
    nome_prestador_fonte = d_nome(bruto["nome_prest"])
    tipo_prestador = texto_celula(bruto["tipo_prest"])
    if not nome or not nasc or not nome_prestador_fonte or not tipo_prestador:
        raise ErroPipeline(
            f"{arquivo.name}, linha {numero_linha}: chave obrigatória ausente"
        )

    regra = MAPA_CODIGOS[codigo]
    if codigo == COD_AT and nome_prestador_fonte != PEDIAKIDS:
        raise ErroPipeline(
            f"{arquivo.name}, linha {numero_linha}: código {COD_AT} só pode ser "
            f"incorporado quando o prestador é {PEDIAKIDS}"
        )

    tipo_normalizado = d_nome(tipo_prestador)
    canal = canal_do(tipo_prestador, nome_prestador_fonte)
    # Guarda de coerência do contrato v5: um registro da Casa jamais pode sair
    # em outro canal (colisão com CREDENCIADA_OVERRIDE). Hoje é impossível nos
    # dados; se um dia acontecer, o pipeline para em vez de decidir sozinho.
    if tipo_normalizado == TIPO_CASA_UNIMED and canal != CANAL_CASA_UNIMED:
        raise ErroPipeline(
            f"{arquivo.name}, linha {numero_linha}: tipo {TIPO_CASA_UNIMED} "
            f"com canal divergente ({canal})"
        )

    registro = {
        "anomes": anomes,
        "matricula": texto_celula(bruto["matricula"]),
        "nome": nome,
        "nasc": nasc,
        "genero": d_nome(bruto["genero"]),
        "guia": texto_celula(bruto["guia"]),
        "dt_atend": texto_celula(bruto["dt_atend"]),
        "cod_sol": texto_celula(bruto["cod_sol"]),
        "nome_sol": d_nome(bruto["nome_sol"]),
        "cbos": texto_celula(bruto["cbos"]),
        "cod_proc": codigo,
        "terapia_id": regra["id"],
        "nome_prest_fonte": nome_prestador_fonte,
        "nome_prest": prestador_consolidado(nome_prestador_fonte),
        "tipo_prest": tipo_normalizado,
        "canal": canal,
        # Entidade publicável, derivada por registro. tipo_prest, nome_prest e
        # nome_prest_fonte permanecem crus: a Matriz 2025 e os contadores
        # 161/160 dependem deles.
        "prest_entidade": (
            CASA_UNIMED
            if tipo_normalizado == TIPO_CASA_UNIMED
            else prestador_consolidado(nome_prestador_fonte)
        ),
        "qtd": qtd,
        "valor": valor,
        "arquivo": arquivo.name,
        "linha": numero_linha,
    }
    registro["paciente"] = chave_paciente(registro)
    return registro


def ler_fonte_xlsx(caminho):
    try:
        import openpyxl
    except ImportError as exc:
        raise ErroPipeline("openpyxl não instalado") from exc

    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    try:
        ws, linha_cabecalho = localizar_aba_dados(wb, caminho)
        registros = []
        for numero_linha, row in enumerate(
            ws.iter_rows(min_row=linha_cabecalho + 1, values_only=True),
            start=linha_cabecalho + 1,
        ):
            valores = list(row)
            if not any(v is not None and texto_celula(v) != "" for v in valores):
                continue
            if len(valores) > len(CABECALHO_FONTE) and any(
                texto_celula(v) for v in valores[len(CABECALHO_FONTE):]
            ):
                raise ErroPipeline(
                    f"{caminho.name}, linha {numero_linha}: dados além das 19 colunas"
                )
            valores = (valores + [None] * len(CABECALHO_FONTE))[:len(CABECALHO_FONTE)]
            registros.append(normalizar_registro(valores, caminho, numero_linha))
    finally:
        wb.close()

    if not registros:
        raise ErroPipeline(f"{caminho.name}: nenhuma linha de dados")
    codigos = {r["cod_proc"] for r in registros}
    if len(codigos) != 1:
        raise ErroPipeline(
            f"{caminho.name}: cada fonte deve conter um único código; encontrados {sorted(codigos)}"
        )
    return registros, next(iter(codigos))


def descobrir_fontes(diretorio, matriz=None):
    raiz = exigir_fora_de_git(diretorio, "diretório de fontes")
    if not raiz.is_dir():
        raise ErroPipeline(f"entrada deve ser um diretório: {raiz}")
    matriz_real = Path(matriz).resolve(strict=True) if matriz else None
    fontes = sorted(
        (
            p.resolve(strict=True)
            for p in raiz.iterdir()
            if p.is_file()
            and p.suffix.lower() == ".xlsx"
            and not p.name.startswith("~$")
            and (matriz_real is None or p.resolve(strict=True) != matriz_real)
        ),
        key=lambda p: p.name.casefold(),
    )
    if len(fontes) != 19:
        raise ErroPipeline(
            f"esperadas exatamente 19 fontes xlsx em {raiz}; encontradas {len(fontes)}"
        )
    return raiz, fontes


def ler_diretorio_fontes(diretorio, matriz=None):
    _, fontes = descobrir_fontes(diretorio, matriz)
    todos = []
    manifesto = []
    codigo_por_arquivo = {}
    for fonte in fontes:
        registros, codigo = ler_fonte_xlsx(fonte)
        if codigo in codigo_por_arquivo:
            raise ErroPipeline(
                f"código {codigo} aparece em duas fontes: "
                f"{codigo_por_arquivo[codigo]} e {fonte.name}"
            )
        codigo_por_arquivo[codigo] = fonte.name
        todos.extend(registros)
        meses = sorted({r["anomes"] for r in registros})
        dentro = [r for r in registros if r["anomes"] in MESES_CANONICOS_SET]
        manifesto.append(
            {
                "arquivo": fonte.name,
                "sha256": sha256_arquivo(fonte),
                "codigo": codigo,
                "terapia_id": MAPA_CODIGOS[codigo]["id"],
                "registros": len(registros),
                "registros_periodo_canonico": len(dentro),
                "competencias": meses,
            }
        )

    recebidos = set(codigo_por_arquivo)
    esperados = set(MAPA_CODIGOS)
    if recebidos != esperados:
        raise ErroPipeline(
            "conjunto de códigos das 19 fontes divergente; "
            f"faltantes={sorted(esperados - recebidos)}, extras={sorted(recebidos - esperados)}"
        )
    return todos, manifesto


def preparar_universo(registros):
    canonicos = [r for r in registros if r["anomes"] in MESES_CANONICOS_SET]
    cobertura = defaultdict(set)
    por_terapia = defaultdict(list)
    for r in canonicos:
        cobertura[r["terapia_id"]].add(r["anomes"])
        por_terapia[r["terapia_id"]].append(r)

    descartadas = []
    mantidas = []
    for terapia_id, definicao in sorted(CATALOGO_TERAPIAS.items()):
        meses = cobertura.get(terapia_id, set())
        faltantes = sorted(MESES_CANONICOS_SET - meses)
        item = {
            "id": terapia_id,
            "terapia": definicao["terapia"],
            "codigos": list(definicao["codigos"]),
            "competencias_encontradas": len(meses),
            "competencias_esperadas": len(MESES_CANONICOS),
            "competencias_faltantes": faltantes,
        }
        if faltantes:
            item["motivo"] = (
                f"cobertura incompleta ({len(meses)}/{len(MESES_CANONICOS)} competências)"
            )
            item["pacientes"] = len(
                {r["paciente"] for r in por_terapia.get(terapia_id, [])}
            )
            descartadas.append(item)
        else:
            mantidas.append(terapia_id)

    if {x["id"] for x in descartadas} != {ID_PECS}:
        raise ErroPipeline(
            "normalização de cobertura deve descartar somente PECS — Fonoaudiologia; "
            f"descartadas={[x['id'] for x in descartadas]}"
        )
    if cobertura[ID_PECS] != MESES_PECS_ESPERADOS:
        raise ErroPipeline(
            "cobertura observada de PECS diverge das 13 competências auditadas"
        )
    if len(mantidas) != 17:
        raise ErroPipeline(f"esperadas 17 terapias mantidas; encontradas {len(mantidas)}")

    incluidos = [r for r in canonicos if r["terapia_id"] in set(mantidas)]
    return incluidos, mantidas, descartadas


def solicitante_valido(registro):
    nome = registro["nome_sol"]
    return bool(
        nome
        and nome != "-"
        and "NAO INFORMADO" not in nome
    )


# As seções seguintes agregam somente conjuntos de pacientes. A chave interna
# Nome + Nascimento existe apenas em memória e nunca é transferida para o JSON.


def gerar_periodos():
    periodos = [
        {
            "id": "tudo",
            "tipo": "tudo",
            "rotulo": "Período completo",
            "inicio": INICIO_CANONICO,
            "fim": FIM_CANONICO,
            "meses": list(MESES_CANONICOS),
        }
    ]
    anos = sorted({m[:4] for m in MESES_CANONICOS})
    for ano in anos:
        meses = [m for m in MESES_CANONICOS if m.startswith(ano)]
        periodos.append(
            {
                "id": f"ano-{ano}",
                "tipo": "ano",
                "rotulo": ano,
                "inicio": meses[0],
                "fim": meses[-1],
                "meses": meses,
            }
        )
    for ano in anos:
        for semestre in (1, 2):
            meses = [
                m for m in MESES_CANONICOS
                if m.startswith(ano)
                and ((int(m[5:7]) - 1) // 6 + 1) == semestre
            ]
            if meses:
                periodos.append(
                    {
                        "id": f"semestre-{ano}-S{semestre}",
                        "tipo": "semestre",
                        "rotulo": f"{semestre}º semestre de {ano}",
                        "inicio": meses[0],
                        "fim": meses[-1],
                        "meses": meses,
                    }
                )
    for ano in anos:
        for trimestre in (1, 2, 3, 4):
            meses = [
                m for m in MESES_CANONICOS
                if m.startswith(ano)
                and ((int(m[5:7]) - 1) // 3 + 1) == trimestre
            ]
            if meses:
                periodos.append(
                    {
                        "id": f"trimestre-{ano}-T{trimestre}",
                        "tipo": "trimestre",
                        "rotulo": f"{trimestre}º trimestre de {ano}",
                        "inicio": meses[0],
                        "fim": meses[-1],
                        "meses": meses,
                    }
                )
    for mes in MESES_CANONICOS:
        periodo_id = mes.replace("/", "-")
        periodos.append(
            {
                "id": f"mes-{periodo_id}",
                "tipo": "mes",
                "rotulo": mes,
                "inicio": mes,
                "fim": mes,
                "meses": [mes],
            }
        )
    return periodos


def metricas_exatas(registros):
    pacientes = {r["paciente"] for r in registros}
    sessoes = sum((r["qtd"] for r in registros), Decimal(0))
    # O PRD exclui somente lançamentos com pagamento igual a zero. Eventuais
    # estornos negativos permanecem no denominador por serem linhas pagas.
    sessoes_pagas = sum(
        (r["qtd"] for r in registros if r["valor"] != 0), Decimal(0)
    )
    pagamento = sum((r["valor"] for r in registros), Decimal(0))
    return {
        "n_pacientes": len(pacientes),
        "pacientes_set": pacientes,
        "registros": len(registros),
        "sessoes": sessoes,
        "sessoes_pagas": sessoes_pagas,
        "pagamento": pagamento,
    }


def publicar_metricas(exatas):
    n = exatas["n_pacientes"]
    if n == 0:
        return {
            "pacientes": 0,
            "registros": 0,
            "sessoes": 0,
            "sessoes_pagas": 0,
            "pagamento": 0.0,
            "custo_sessao": None,
            "custo_beneficiario": None,
            "sessoes_beneficiario": None,
        }
    return {
        "pacientes": n,
        "registros": exatas["registros"],
        "sessoes": numero(exatas["sessoes"]),
        "sessoes_pagas": numero(exatas["sessoes_pagas"]),
        "pagamento": dinheiro(exatas["pagamento"]),
        "custo_sessao": razao(exatas["pagamento"], exatas["sessoes_pagas"]),
        "custo_beneficiario": razao(exatas["pagamento"], n),
        "sessoes_beneficiario": razao(exatas["sessoes"], n),
    }


def grupos(registros, chave):
    saida = defaultdict(list)
    for registro in registros:
        saida[chave(registro)].append(registro)
    return saida


def pct(parte, total):
    if not total:
        return 0.0
    return dinheiro(Decimal(parte) * Decimal(100) / Decimal(total))


def faixa_etaria(idade):
    if idade < 3:
        return "0–2"
    if idade < 6:
        return "3–5"
    if idade < 9:
        return "6–8"
    if idade < 12:
        return "9–11"
    if idade < 15:
        return "12–14"
    if idade < 18:
        return "15–17"
    return "18+"


def idade_na_data(nascimento, corte):
    nasc = date.fromisoformat(nascimento)
    return corte.year - nasc.year - (
        (corte.month, corte.day) < (nasc.month, nasc.day)
    )


def serie_mensal(registros, meses):
    por_mes = grupos(registros, lambda r: r["anomes"])
    serie = []
    for mes in meses:
        item = {"m": mes}
        item.update(publicar_metricas(metricas_exatas(por_mes.get(mes, []))))
        serie.append(item)
    return serie


def agregar_terapias(
    registros,
    terapias_mantidas,
    pagamento_total,
):
    por_terapia = grupos(registros, lambda r: r["terapia_id"])
    itens = []
    for terapia_id in terapias_mantidas:
        linhas = por_terapia.get(terapia_id, [])
        exatas = metricas_exatas(linhas)
        definicao = CATALOGO_TERAPIAS[terapia_id]
        item = {
            "id": terapia_id,
            "terapia": definicao["terapia"],
            "codigos": list(definicao["codigos"]),
            "especialidade": definicao["especialidade"],
            "metodo": definicao["metodo"],
            "quantidade_entidades": 1,
        }
        item.update(publicar_metricas(exatas))
        item["exposicoes"] = len(
            {(r["paciente"], r["terapia_id"]) for r in linhas}
        )
        item["participacao_pct"] = pct(
            exatas["pagamento"], pagamento_total
        )
        at = metricas_exatas(
            [r for r in linhas if r["cod_proc"] == COD_AT]
        )
        item["at_sessoes"] = numero(at["sessoes"])
        item["at_pct"] = (
            pct(at["sessoes"], exatas["sessoes"])
            if exatas["sessoes"] else None
        )
        itens.append(item)
    itens.sort(key=lambda x: x["terapia"])
    return itens


def agregar_prestadores(registros, meses):
    # Contrato v5: ranking por entidade publicada. Os registros da Casa Unimed
    # consolidam num item único com a composição por profissional.
    por_prestador = grupos(registros, lambda r: r["prest_entidade"])
    itens = []
    series = {}
    for nome, linhas in por_prestador.items():
        exatas = metricas_exatas(linhas)
        item = {
            "nome": nome,
            "quantidade_entidades": 1,
        }
        item.update(publicar_metricas(exatas))
        paciente_meses = len(
            {(r["paciente"], r["anomes"]) for r in linhas}
        )
        item["paciente_meses"] = paciente_meses
        item["intensidade_mensal"] = (
            razao(exatas["sessoes"], paciente_meses)
            if paciente_meses else None
        )
        item["nomes_fonte"] = sorted(
            {r["nome_prest_fonte"] for r in linhas}
        )
        canal_pagamento = defaultdict(Decimal)
        canal_registros = Counter()
        for r in linhas:
            canal_pagamento[r["canal"]] += r["valor"]
            canal_registros[r["canal"]] += 1
        item["canais"] = sorted(canal_pagamento)
        item["canal"] = max(
            canal_pagamento,
            key=lambda c: (canal_pagamento[c], canal_registros[c], c),
        )
        terapia_ids = sorted({r["terapia_id"] for r in linhas})
        item["terapias"] = [
            CATALOGO_TERAPIAS[t]["terapia"] for t in terapia_ids
        ]
        item["especialidades"] = sorted(
            {CATALOGO_TERAPIAS[t]["especialidade"] for t in terapia_ids}
        )
        item["metodos"] = sorted(
            {CATALOGO_TERAPIAS[t]["metodo"] for t in terapia_ids}
        )
        at = metricas_exatas(
            [r for r in linhas if r["cod_proc"] == COD_AT]
        )
        item["at_sessoes"] = numero(at["sessoes"])
        item["at_pct"] = (
            pct(at["sessoes"], exatas["sessoes"])
            if exatas["sessoes"] else None
        )
        if nome == CASA_UNIMED:
            composicao = []
            for profissional, plinhas in grupos(
                linhas, lambda r: r["nome_prest"]
            ).items():
                pexatas = metricas_exatas(plinhas)
                composicao.append(
                    {
                        "nome": profissional,
                        "sessoes": numero(pexatas["sessoes"]),
                        "pagamento": dinheiro(pexatas["pagamento"]),
                        "pacientes": pexatas["n_pacientes"],
                    }
                )
            composicao.sort(key=lambda x: (-(x["pagamento"] or 0), x["nome"]))
            item["composicao"] = composicao
            item["quantidade_entidades"] = len(composicao)
        itens.append(item)
        series[nome] = serie_mensal(linhas, meses)

    itens.sort(key=lambda x: (-(x["pagamento"] or 0), x["nome"]))
    return itens, dict(sorted(series.items()))


def agregar_prestador_disciplina(registros, chave_nome=None):
    # Contrato v4: fatia exata prestador × disciplina. A soma das fatias de um
    # prestador reproduz o total dele no recorte (validado em validar_saida).
    # Contrato v5: a chave padrão é a entidade publicada (coerente com o
    # ranking); as referências continuam consumindo fatias por profissional
    # via chave_nome=lambda r: r["nome_prest"].
    if chave_nome is None:
        chave_nome = lambda r: r["prest_entidade"]
    por_fatia = grupos(
        registros,
        lambda r: (
            chave_nome(r),
            CATALOGO_TERAPIAS[r["terapia_id"]]["especialidade"],
        ),
    )
    itens = []
    for (nome, disciplina), linhas in por_fatia.items():
        exatas = metricas_exatas(linhas)
        item = {"nome": nome, "disciplina": disciplina}
        item.update(publicar_metricas(exatas))
        paciente_meses = len({(r["paciente"], r["anomes"]) for r in linhas})
        item["paciente_meses"] = paciente_meses
        item["intensidade_mensal"] = (
            razao(exatas["sessoes"], paciente_meses) if paciente_meses else None
        )
        canal_pagamento = defaultdict(Decimal)
        canal_registros = Counter()
        for r in linhas:
            canal_pagamento[r["canal"]] += r["valor"]
            canal_registros[r["canal"]] += 1
        item["canal"] = max(
            canal_pagamento,
            key=lambda c: (canal_pagamento[c], canal_registros[c], c),
        )
        at = metricas_exatas([r for r in linhas if r["cod_proc"] == COD_AT])
        item["at_sessoes"] = numero(at["sessoes"])
        item["at_pct"] = (
            pct(at["sessoes"], exatas["sessoes"]) if exatas["sessoes"] else None
        )
        itens.append(item)
    itens.sort(key=lambda x: (x["nome"], x["disciplina"]))
    return itens


def disciplina_do(registro):
    return CATALOGO_TERAPIAS[registro["terapia_id"]]["especialidade"]


def agregar_disciplina_serie(registros, meses):
    # Contrato v4: série mensal por disciplina. A soma das disciplinas em cada
    # competência reproduz a série mensal total (validado em validar_saida).
    por_disciplina = grupos(
        registros,
        lambda r: CATALOGO_TERAPIAS[r["terapia_id"]]["especialidade"],
    )
    return {
        disciplina: serie_mensal(linhas, meses)
        for disciplina, linhas in sorted(por_disciplina.items())
    }


def agregar_disciplina_canais(registros):
    # Contrato v5: leitura por canal dentro de cada disciplina, para a aba
    # Canais reagir ao escopo de disciplina. Cross-foot duplo validado em
    # validar_saida: contra prestador_disciplina e contra canais do recorte.
    por_disciplina = grupos(registros, disciplina_do)
    saida = {}
    for disciplina, linhas in sorted(por_disciplina.items()):
        itens = []
        for canal, linhas_canal in sorted(
            grupos(linhas, lambda r: r["canal"]).items()
        ):
            item = {"canal": canal}
            item.update(publicar_metricas(metricas_exatas(linhas_canal)))
            item["exposicoes"] = len(
                {(r["paciente"], r["terapia_id"]) for r in linhas_canal}
            )
            item["prestadores"] = len(
                {r["prest_entidade"] for r in linhas_canal}
            )
            item["profissionais"] = len(
                {r["nome_prest"] for r in linhas_canal}
            )
            itens.append(item)
        itens.sort(key=lambda x: (-(x["pagamento"] or 0), x["canal"]))
        saida[disciplina] = itens
    return saida


def agregar_resumo_disciplina(registros):
    # Contrato v5: resumo por disciplina com crianças únicas deduplicadas
    # DENTRO da disciplina. A soma de pacientes entre disciplinas excede o
    # total do recorte por construção; pagamento e sessões somam exatos.
    por_disciplina = grupos(registros, disciplina_do)
    saida = {}
    for disciplina, linhas in sorted(por_disciplina.items()):
        item = publicar_metricas(metricas_exatas(linhas))
        item["exposicoes"] = len(
            {(r["paciente"], r["terapia_id"]) for r in linhas}
        )
        saida[disciplina] = item
    return saida


def kit_da_crianca(disciplinas):
    # Taxonomia do painel: classificação exaustiva e mutuamente exclusiva da
    # criança pelo conjunto de disciplinas utilizadas no recorte.
    if len(disciplinas) == 1:
        return "unica"
    if len(disciplinas) == 2:
        return "duas"
    if disciplinas == TRIPE_DISCIPLINAS:
        return "tripe"
    if TRIPE_DISCIPLINAS <= disciplinas:
        return "tripe-ampliado"
    return "multiplas-sem-tripe"


def agregar_kits(registros):
    # Contrato v5: agregados por kit terapêutico. Kit ausente do JSON significa
    # suprimido pelo k-anonimato (menos de K_MINIMO_KIT crianças), nunca zero.
    disciplinas_paciente = defaultdict(set)
    for r in registros:
        disciplinas_paciente[r["paciente"]].add(disciplina_do(r))
    linhas_kit = grupos(
        registros, lambda r: kit_da_crianca(disciplinas_paciente[r["paciente"]])
    )
    saida = {}
    for kit in KITS:
        linhas = linhas_kit.get(kit["id"], [])
        criancas = {r["paciente"] for r in linhas}
        if len(criancas) < K_MINIMO_KIT:
            continue
        exatas = metricas_exatas(linhas)
        item = {
            "id": kit["id"],
            "rotulo": kit["rotulo"],
            "criterio": kit["criterio"],
            "criancas": len(criancas),
        }
        item.update(publicar_metricas(exatas))
        crianca_meses = len({(r["paciente"], r["anomes"]) for r in linhas})
        item["crianca_meses"] = crianca_meses
        item["custo_crianca_mes"] = razao(exatas["pagamento"], crianca_meses)
        canais = []
        for canal, linhas_canal in sorted(
            grupos(linhas, lambda r: r["canal"]).items()
        ):
            canal_exatas = metricas_exatas(linhas_canal)
            canais.append(
                {
                    "canal": canal,
                    "pagamento": dinheiro(canal_exatas["pagamento"]),
                    "sessoes": numero(canal_exatas["sessoes"]),
                    "pct_pagamento": pct(
                        canal_exatas["pagamento"], exatas["pagamento"]
                    ),
                }
            )
        canais.sort(key=lambda x: (-(x["pagamento"] or 0), x["canal"]))
        item["canais"] = canais
        saida[kit["id"]] = item
    return saida


def referencias_disciplina(fatias):
    # Contrato v4: mediana e quartis do R$/sessão e da intensidade por
    # disciplina (e por canal dentro dela), sobre fatias com volume mínimo.
    # Células com menos de MIN_PRESTADORES_REFERENCIA prestadores não publicam.
    preco = {}
    intensidade = {}
    por_disciplina = defaultdict(list)
    for item in fatias:
        if (
            item["sessoes"] is not None
            and Decimal(str(item["sessoes"])) >= MIN_SESSOES_REFERENCIA
            and item["custo_sessao"] is not None
        ):
            por_disciplina[item["disciplina"]].append(item)
    def quartis(valores):
        decimais = sorted(Decimal(str(v)) for v in valores)
        return {
            "mediana": dinheiro(percentil_linear(decimais, Decimal("0.5"))),
            "p25": dinheiro(percentil_linear(decimais, Decimal("0.25"))),
            "p75": dinheiro(percentil_linear(decimais, Decimal("0.75"))),
            "n": len(decimais),
        }

    for disciplina, itens in sorted(por_disciplina.items()):
        precos = [x["custo_sessao"] for x in itens]
        if len(precos) >= MIN_PRESTADORES_REFERENCIA:
            preco[disciplina] = quartis(precos)
            preco[disciplina]["por_canal"] = {}
            por_canal = defaultdict(list)
            for x in itens:
                por_canal[x["canal"]].append(x["custo_sessao"])
            for canal, valores in sorted(por_canal.items()):
                if len(valores) >= MIN_PRESTADORES_REFERENCIA:
                    preco[disciplina]["por_canal"][canal] = quartis(valores)
        intensidades = [
            x["intensidade_mensal"]
            for x in itens
            if x["intensidade_mensal"] is not None
        ]
        if len(intensidades) >= MIN_PRESTADORES_REFERENCIA:
            intensidade[disciplina] = quartis(intensidades)
    return {"preco_disciplina": preco, "intensidade_disciplina": intensidade}


def distribuir_linhas_cuidado(registros, total_pacientes, periodo_id):
    terapias_por_paciente = defaultdict(set)
    codigos_por_paciente = defaultdict(set)
    for r in registros:
        terapias_por_paciente[r["paciente"]].add(r["terapia_id"])
        codigos_por_paciente[r["paciente"]].add(r["cod_proc"])

    dist_clinica = Counter(len(v) for v in terapias_por_paciente.values())
    dist_componentes = Counter(len(v) for v in codigos_por_paciente.values())
    combinacoes = Counter(
        tuple(sorted(v)) for v in terapias_por_paciente.values()
    )

    distribuicao_exata = [
        {
            "faixa": (
                "1 terapia" if quantidade == 1 else f"{quantidade} terapias"
            ),
            "quantidade_terapias": quantidade,
            "pacientes": pacientes,
            "pct": pct(pacientes, total_pacientes),
        }
        for quantidade, pacientes in sorted(dist_clinica.items())
    ]

    combinacoes_exatas = [
        {
            "rotulo": " + ".join(
                CATALOGO_TERAPIAS[x]["terapia"] for x in ids
            ),
            "terapias": [
                CATALOGO_TERAPIAS[x]["terapia"] for x in ids
            ],
            "pacientes": n,
            "pct": pct(n, total_pacientes),
            "quantidade_combinacoes": 1,
        }
        for ids, n in sorted(
            combinacoes.items(), key=lambda item: (-item[1], item[0])
        )
    ]

    exposicoes_clinicas = sum(
        quantidade * pacientes
        for quantidade, pacientes in dist_clinica.items()
    )
    exposicoes_componentes = sum(
        quantidade * pacientes
        for quantidade, pacientes in dist_componentes.items()
    )
    dois_mais_clinico = sum(
        n for quantidade, n in dist_clinica.items() if quantidade >= 2
    )
    dois_mais_componentes = sum(
        n for quantidade, n in dist_componentes.items() if quantidade >= 2
    )
    quatro_mais_clinico = sum(
        n for quantidade, n in dist_clinica.items() if quantidade >= 4
    )
    quatro_mais_componentes = sum(
        n for quantidade, n in dist_componentes.items() if quantidade >= 4
    )
    reconciliacao_componentes = None
    if periodo_id == "tudo":
        reconciliacao_componentes = {
            "natureza": (
                "Reconciliação técnica por códigos assistenciais; o código de "
                "assistente terapêutica integra ABA — Psicologia e não constitui "
                "uma terapia clínica adicional."
            ),
            "distribuicao": [
                {
                    "faixa": (
                        "1 componente"
                        if quantidade == 1
                        else f"{quantidade} componentes"
                    ),
                    "quantidade_componentes": quantidade,
                    "pacientes": pacientes,
                    "pct": pct(pacientes, total_pacientes),
                }
                for quantidade, pacientes in sorted(dist_componentes.items())
            ],
            "exposicoes": exposicoes_componentes,
            "dois_ou_mais_pct": pct(
                dois_mais_componentes, total_pacientes
            ),
            "quatro_ou_mais_pct": pct(
                quatro_mais_componentes, total_pacientes
            ),
        }

    completo = {
        "distribuicao": distribuicao_exata,
        "exposicoes": exposicoes_clinicas,
        "dois_ou_mais_pct": pct(dois_mais_clinico, total_pacientes),
        "quatro_ou_mais_pct": pct(quatro_mais_clinico, total_pacientes),
        "combinacoes": combinacoes_exatas,
        "pacientes_em_combinacoes_agrupadas": 0,
        "reconciliacao_componentes_codigo": reconciliacao_componentes,
    }
    auditoria = {
        "distribuicao_clinica": dict(sorted(dist_clinica.items())),
        "distribuicao_componentes": dict(sorted(dist_componentes.items())),
        "exposicoes_clinicas": exposicoes_clinicas,
        "exposicoes_componentes": exposicoes_componentes,
        "dois_mais_clinico": dois_mais_clinico,
        "dois_mais_componentes": dois_mais_componentes,
        "quatro_mais_clinico": quatro_mais_clinico,
        "quatro_mais_componentes": quatro_mais_componentes,
        "combinacoes_exatas": dict(combinacoes),
        "combinacoes_publicadas": {
            item["rotulo"]: item["pacientes"]
            for item in combinacoes_exatas
        },
    }
    return completo, auditoria


def agregar_canais(registros):
    por_canal = grupos(registros, lambda r: r["canal"])
    itens = []
    for canal, linhas in por_canal.items():
        item = {"canal": canal}
        item.update(publicar_metricas(metricas_exatas(linhas)))
        # Contrato v5: prestadores conta entidades publicadas (a Casa Unimed
        # vale 1); profissionais conta pessoas físicas/CNPJs de origem.
        item["prestadores"] = len({r["prest_entidade"] for r in linhas})
        item["profissionais"] = len({r["nome_prest"] for r in linhas})
        itens.append(item)
    itens.sort(
        key=lambda x: (
            x["pagamento"] is None,
            -(x["pagamento"] or 0),
            x["canal"],
        )
    )
    return itens


def agregar_solicitantes(registros):
    elegiveis = [r for r in registros if solicitante_valido(r)]
    por_solicitante = grupos(elegiveis, lambda r: r["nome_sol"])
    itens = []
    for nome, linhas in por_solicitante.items():
        exatas = metricas_exatas(linhas)
        item = {"nome": nome}
        item.update(publicar_metricas(exatas))
        item["cbos"] = Counter(r["cbos"] for r in linhas).most_common(1)[0][0]
        executores = []
        # Contrato v5: na visão do solicitante o encaminhamento é para a
        # entidade publicada — os profissionais da Casa consolidam nela.
        for prestador, par in grupos(linhas, lambda r: r["prest_entidade"]).items():
            par_exato = metricas_exatas(par)
            executante = {"nome": prestador}
            executante.update(publicar_metricas(par_exato))
            executante["pct_pagamento"] = pct(
                par_exato["pagamento"], exatas["pagamento"]
            )
            executores.append(executante)
        executores.sort(
            key=lambda x: (-(x["pagamento"] or 0), x["nome"])
        )
        principal = executores[0] if executores else None
        item["principal"] = principal["nome"] if principal else None
        item["principal_pct"] = principal["pct_pagamento"] if principal else None
        item["executores"] = executores
        itens.append(item)
    itens.sort(key=lambda x: (-(x["pagamento"] or 0), x["nome"]))
    return {"itens": itens}


def agregar_pacientes(registros, corte):
    por_paciente = grupos(registros, lambda r: r["paciente"])
    faixa_pacientes = defaultdict(set)
    genero_pacientes = defaultdict(set)
    intensidade = {}
    custo_paciente = {}
    for paciente, linhas in por_paciente.items():
        idade = idade_na_data(paciente[1], corte)
        faixa_pacientes[faixa_etaria(idade)].add(paciente)
        genero_pacientes[linhas[0]["genero"] or "NAO INFORMADO"].add(paciente)
        intensidade[paciente] = sum((r["qtd"] for r in linhas), Decimal(0))
        custo_paciente[paciente] = sum((r["valor"] for r in linhas), Decimal(0))

    def metricas_conjunto(pacientes):
        return publicar_metricas(
            metricas_exatas(
                [r for r in registros if r["paciente"] in pacientes]
            )
        )

    faixas = []
    for faixa in FAIXAS_ETARIAS:
        item = {"faixa": faixa}
        item.update(metricas_conjunto(faixa_pacientes.get(faixa, set())))
        faixas.append(item)

    genero = []
    for rotulo, pacientes in sorted(genero_pacientes.items()):
        item = {"genero": rotulo}
        item.update(metricas_conjunto(pacientes))
        genero.append(item)

    bucket_pacientes = defaultdict(set)
    for paciente, sessoes in intensidade.items():
        if sessoes < 3:
            bucket = "<3"
        elif sessoes < 5:
            bucket = "3–4"
        elif sessoes < 9:
            bucket = "5–8"
        else:
            bucket = "9+"
        bucket_pacientes[bucket].add(paciente)
    buckets = [
        {
            "faixa": bucket,
            "pacientes": len(bucket_pacientes.get(bucket, set())),
            "pct": pct(
                len(bucket_pacientes.get(bucket, set())), len(por_paciente)
            ),
        }
        for bucket in BUCKETS_INTENSIDADE
    ]

    valores_intensidade = list(intensidade.values())
    valores_custo = list(custo_paciente.values())
    if por_paciente:
        media_intensidade = dinheiro(statistics.mean(valores_intensidade))
        mediana_intensidade = dinheiro(statistics.median(valores_intensidade))
        media_custo = dinheiro(statistics.mean(valores_custo))
        mediana_custo = dinheiro(statistics.median(valores_custo))
    else:
        media_intensidade = mediana_intensidade = None
        media_custo = mediana_custo = None
    return {
        "faixa_etaria": faixas,
        "intensidade": {
            "definicao": "Soma de sessões por beneficiário no período selecionado.",
            "buckets": buckets,
            "media": media_intensidade,
            "mediana": mediana_intensidade,
        },
        "genero": genero,
        "custo_por_paciente": {
            "media": media_custo,
            "mediana": mediana_custo,
        },
    }


def agregar_recorte(registros, periodo, terapias_mantidas, corte):
    meses = set(periodo["meses"])
    linhas = [r for r in registros if r["anomes"] in meses]
    exatas = metricas_exatas(linhas)
    resumo = publicar_metricas(exatas)
    resumo["exposicoes"] = len(
        {(r["paciente"], r["terapia_id"]) for r in linhas}
    )
    resumo["componentes_codigo"] = len(
        {(r["paciente"], r["cod_proc"]) for r in linhas}
    )
    resumo["prestadores_observados"] = len(
        {r["nome_prest_fonte"] for r in linhas}
    )
    resumo["prestadores_consolidados"] = len(
        {r["nome_prest"] for r in linhas}
    )
    # Contrato v5: contadores da camada de entidades publicadas. Os dois
    # campos acima permanecem intactos — ancoram a auditoria (161/160).
    profissionais_casa = {
        r["nome_prest"] for r in linhas if r["prest_entidade"] == CASA_UNIMED
    }
    profissionais_fora = {
        r["nome_prest"] for r in linhas if r["prest_entidade"] != CASA_UNIMED
    }
    resumo["prestadores_publicados"] = len(
        {r["prest_entidade"] for r in linhas}
    )
    resumo["profissionais_casa_unimed"] = len(profissionais_casa)
    linhas_cuidado, auditoria_linhas = distribuir_linhas_cuidado(
        linhas, exatas["n_pacientes"], periodo["id"]
    )
    prestadores, serie_prestador = agregar_prestadores(
        linhas, periodo["meses"]
    )
    fatias_disciplina = agregar_prestador_disciplina(linhas)
    # Referências permanecem na granularidade profissional (nome_prest):
    # colapsar dezenas de profissionais da Casa em uma observação distorceria
    # mediana, quartis e o piso k>=5. As fatias abaixo não são publicadas.
    fatias_profissional = agregar_prestador_disciplina(
        linhas, chave_nome=lambda r: r["nome_prest"]
    )
    por_disciplina = grupos(linhas, disciplina_do)
    publico = {
        "resumo": resumo,
        "serie_mensal": serie_mensal(linhas, periodo["meses"]),
        "terapias": agregar_terapias(
            linhas, terapias_mantidas, exatas["pagamento"]
        ),
        "prestadores": prestadores,
        "serie_prestador": serie_prestador,
        "prestador_disciplina": fatias_disciplina,
        "disciplina_serie": agregar_disciplina_serie(linhas, periodo["meses"]),
        "disciplina_canais": agregar_disciplina_canais(linhas),
        "resumo_disciplina": agregar_resumo_disciplina(linhas),
        "pacientes_disciplina": {
            disciplina: agregar_pacientes(linhas_disc, corte)
            for disciplina, linhas_disc in sorted(por_disciplina.items())
        },
        "kits": agregar_kits(linhas),
        "referencias": referencias_disciplina(fatias_profissional),
        "linhas_cuidado": linhas_cuidado,
        "canais": agregar_canais(linhas),
        "solicitantes": agregar_solicitantes(linhas),
        "pacientes": agregar_pacientes(linhas, corte),
    }
    auditoria = {
        "metricas": exatas,
        "linhas_cuidado": auditoria_linhas,
        "prestadores_observados": resumo["prestadores_observados"],
        "prestadores_consolidados": resumo["prestadores_consolidados"],
        "prestadores_publicados": resumo["prestadores_publicados"],
        "profissionais_exclusivos_casa": len(
            profissionais_casa - profissionais_fora
        ),
    }
    return publico, auditoria


def construir_dados(
    registros_todos,
    registros_incluidos,
    manifesto,
    terapias_mantidas,
    terapias_descartadas,
    corte,
):
    periodos = gerar_periodos()
    recortes = {}
    auditorias = {}
    for periodo in periodos:
        recorte, auditoria = agregar_recorte(
            registros_incluidos, periodo, terapias_mantidas, corte
        )
        recortes[periodo["id"]] = recorte
        auditorias[periodo["id"]] = auditoria

    catalogo_publico = [
        CATALOGO_TERAPIAS[terapia_id]
        for terapia_id in sorted(
            terapias_mantidas,
            key=lambda x: CATALOGO_TERAPIAS[x]["terapia"],
        )
    ]
    at_exato = metricas_exatas(
        [r for r in registros_incluidos if r["cod_proc"] == COD_AT]
    )
    at_publico = publicar_metricas(at_exato)
    contagens = dict(recortes["tudo"]["resumo"])
    contagens.update(
        {
            "terapias": len(terapias_mantidas),
            "fontes": len(manifesto),
        }
    )
    dados = {
        "meta": {
            "painel": "Painel Terapias Especiais — Unimed Governador Valadares",
            "versao_script": VERSAO,
            "gerado_em": date.today().isoformat(),
            "data_corte_idade": corte.isoformat(),
            "classificacao": "agregados completos de uso interno",
            "identificadores_beneficiario": "não incluídos",
            "periodo_canonico": {
                "inicio": INICIO_CANONICO,
                "fim": FIM_CANONICO,
                "competencias": len(MESES_CANONICOS),
            },
            "periodos": periodos,
            "terapias": catalogo_publico,
            "terapias_descartadas": terapias_descartadas,
            "contagens": contagens,
            "manifesto_fontes": manifesto,
            "excecao_at": {
                "codigo": COD_AT,
                "prestador": PEDIAKIDS,
                "terapia_destino": "Terapia ABA — Psicologia",
                "origem_preservada": True,
                **at_publico,
            },
            "casa_unimed": {
                "entidade": CASA_UNIMED,
                "criterio": f"registro com Tipo Prestador = {TIPO_CASA_UNIMED}",
                "consolidacao": (
                    "por registro; o mesmo profissional pode figurar como "
                    "prestador individual nos demais canais"
                ),
                "profissionais": contagens["profissionais_casa_unimed"],
            },
            "kits_catalogo": {
                "regra": (
                    "classificação por criança pelo conjunto de disciplinas "
                    "utilizadas no recorte; kit ausente de um recorte foi "
                    f"suprimido pelo k-anonimato (menos de {K_MINIMO_KIT} "
                    "crianças), nunca é zero"
                ),
                "k_minimo": K_MINIMO_KIT,
                "kits": [dict(kit) for kit in KITS],
            },
            "referencias_granularidade": "profissional (nome_prest)",
            "matriz_2025": None,
        },
        "recortes": recortes,
    }
    auditoria = {
        "recortes": auditorias,
        "registros_todos": registros_todos,
        "registros_incluidos": registros_incluidos,
    }
    return dados, auditoria


# ---------------- reconciliação com a Matriz Analítica 2025 ----------------


def percentil_linear(valores, proporcao):
    ordenados = sorted(valores)
    if not ordenados:
        return None
    posicao = (len(ordenados) - 1) * Decimal(str(proporcao))
    inferior = int(math.floor(posicao))
    superior = int(math.ceil(posicao))
    if inferior == superior:
        return ordenados[inferior]
    fracao = posicao - inferior
    return ordenados[inferior] + (
        ordenados[superior] - ordenados[inferior]
    ) * fracao


def estatisticas_intensidade(registros):
    sessoes_paciente = defaultdict(Decimal)
    for r in registros:
        sessoes_paciente[r["paciente"]] += r["qtd"]
    valores = list(sessoes_paciente.values())
    if not valores:
        return {
            "beneficiarios": 0,
            "media": None,
            "mediana": None,
            "p25": None,
            "p75": None,
            "minimo": None,
            "maximo": None,
            "desvio_padrao": None,
        }
    media = statistics.mean(valores)
    return {
        "beneficiarios": len(valores),
        "media": media,
        "mediana": statistics.median(valores),
        "p25": percentil_linear(valores, Decimal("0.25")),
        "p75": percentil_linear(valores, Decimal("0.75")),
        "minimo": min(valores),
        "maximo": max(valores),
        # A Matriz 2025 registra dispersão zero quando há uma única criança no
        # estrato. Mantemos essa convenção para que a conferência seja literal;
        # estratos vazios continuam representados por nulo no retorno acima.
        "desvio_padrao": statistics.stdev(valores) if len(valores) > 1 else Decimal("0"),
    }


def rede_matriz(registro):
    return (
        "Rede própria"
        if registro["tipo_prest"] == "ATENDIMENTO ESPECIALIZADO"
        else "Rede externa"
    )


def agregados_matriz_esperados(registros):
    ano = [r for r in registros if r["anomes"].startswith("2025/")]
    por_terapia = grupos(ano, lambda r: r["terapia_id"])
    resumo = {}
    mensal = {}
    rede = {}

    for terapia_id, linhas in por_terapia.items():
        propria = [r for r in linhas if rede_matriz(r) == "Rede própria"]
        externa = [r for r in linhas if rede_matriz(r) == "Rede externa"]
        total = metricas_exatas(linhas)
        ep = metricas_exatas(propria)
        ee = metricas_exatas(externa)
        ip = estatisticas_intensidade(propria)
        ie = estatisticas_intensidade(externa)
        resumo[terapia_id] = {
            "Registros": total["registros"],
            "Beneficiários*": total["n_pacientes"],
            "Sessões": total["sessoes"],
            "Pagamento": total["pagamento"],
            "Ben. própria": ep["n_pacientes"],
            "Sessões própria": ep["sessoes"],
            "Média própria": ip["media"],
            "Mediana própria": ip["mediana"],
            "Ben. externa": ee["n_pacientes"],
            "Sessões externa": ee["sessoes"],
            "Média externa": ie["media"],
            "Mediana externa": ie["mediana"],
            "Custo/sessão própria": (
                ep["pagamento"] / ep["sessoes_pagas"]
                if ep["sessoes_pagas"] else None
            ),
            "Custo/sessão externa": (
                ee["pagamento"] / ee["sessoes_pagas"]
                if ee["sessoes_pagas"] else None
            ),
        }

        pacientes_propria = {r["paciente"] for r in propria}
        pacientes_externa = {r["paciente"] for r in externa}
        sobreposicao = len(pacientes_propria & pacientes_externa)
        for nome_rede, linhas_rede in grupos(linhas, rede_matriz).items():
            em = metricas_exatas(linhas_rede)
            intensidade = estatisticas_intensidade(linhas_rede)
            rede[(terapia_id, nome_rede)] = {
                "Beneficiários": em["n_pacientes"],
                "Sessões": em["sessoes"],
                "Pagamento": em["pagamento"],
                "Média": intensidade["media"],
                "Mediana": intensidade["mediana"],
                "P25": intensidade["p25"],
                "P75": intensidade["p75"],
                "Mínimo": intensidade["minimo"],
                "Máximo": intensidade["maximo"],
                "Desvio-padrão": intensidade["desvio_padrao"],
                "Custo/sessão": (
                    em["pagamento"] / em["sessoes_pagas"]
                    if em["sessoes_pagas"] else None
                ),
                "Custo/beneficiário": (
                    em["pagamento"] / em["n_pacientes"]
                    if em["n_pacientes"] else None
                ),
                "Sobreposição": sobreposicao,
            }

    for (terapia_id, mes), linhas in grupos(
        ano, lambda r: (r["terapia_id"], r["anomes"])
    ).items():
        propria = [r for r in linhas if rede_matriz(r) == "Rede própria"]
        externa = [r for r in linhas if rede_matriz(r) == "Rede externa"]
        total = metricas_exatas(linhas)
        ep = metricas_exatas(propria)
        ee = metricas_exatas(externa)
        mensal[(terapia_id, mes)] = {
            "Registros": total["registros"],
            "Beneficiários": total["n_pacientes"],
            "Sessões": total["sessoes"],
            "Pagamento": total["pagamento"],
            "Ben. própria": ep["n_pacientes"],
            "Sessões própria": ep["sessoes"],
            "Ben. externa": ee["n_pacientes"],
            "Sessões externa": ee["sessoes"],
        }
    return {
        "Resumo_Executivo": resumo,
        "Mensal": mensal,
        "Rede_por_Terapia": rede,
    }


def localizar_cabecalho_matriz(ws, primeira_coluna):
    for numero_linha, row in enumerate(
        ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1
    ):
        valores = [texto_celula(v) for v in row]
        if valores and valores[0] == primeira_coluna:
            return numero_linha, valores
    raise ErroPipeline(
        f"Matriz, aba {ws.title}: cabeçalho {primeira_coluna!r} não encontrado"
    )


def indice_cabecalho(cabecalho, nome, aba):
    try:
        return cabecalho.index(nome)
    except ValueError as exc:
        raise ErroPipeline(
            f"Matriz, aba {aba}: coluna obrigatória ausente: {nome}"
        ) from exc


def id_terapia_por_codigo_matriz(valor):
    codigos = re.findall(r"\d{8}", texto_celula(valor))
    ids = {MAPA_CODIGOS[c]["id"] for c in codigos if c in MAPA_CODIGOS}
    if not ids:
        return None
    if len(ids) != 1:
        raise ErroPipeline(
            f"Matriz: códigos da mesma linha apontam para terapias distintas: {codigos}"
        )
    return next(iter(ids))


def mapa_nomes_matriz():
    mapa = {
        chave_rotulo(item["terapia"]): terapia_id
        for terapia_id, item in CATALOGO_TERAPIAS.items()
    }
    mapa[chave_rotulo("Terapias especiais cobertas conforme parecer ANS")] = (
        "ans-centros-referencia"
    )
    return mapa


MAPA_NOMES_MATRIZ = mapa_nomes_matriz()


def id_terapia_por_nome_matriz(valor):
    return MAPA_NOMES_MATRIZ.get(chave_rotulo(valor))


def decimal_matriz(valor):
    if valor is None or texto_celula(valor) == "":
        return None
    return decimal_celula(valor, "valor da Matriz")


def valores_equivalentes(observado, esperado, campo):
    if observado is None or esperado is None:
        return observado is None and esperado is None
    od = Decimal(str(observado))
    ed = Decimal(str(esperado))
    if campo in {
        "Registros", "Beneficiários*", "Beneficiários", "Sessões",
        "Ben. própria", "Sessões própria", "Ben. externa",
        "Sessões externa", "Sobreposição",
    }:
        return od == ed
    if "Pagamento" in campo or "Custo/" in campo:
        # Conferência monetária ±0 no centavo publicado. A quantização absorve
        # apenas a representação binária do Excel, nunca diferença de 1 centavo.
        centavo = Decimal("0.01")
        return (
            od.quantize(centavo, rounding=ROUND_HALF_UP)
            == ed.quantize(centavo, rounding=ROUND_HALF_UP)
        )
    return abs(od - ed) <= Decimal("0.000001")


def comparar_celula(controle, aba, chave, campo, observado, esperado):
    controle["campos_comparados"] += 1
    if not valores_equivalentes(observado, esperado, campo):
        controle["divergencias"].append(
            f"{aba} {chave} {campo}: Matriz={observado!r}, pipeline={esperado!r}"
        )


def reconciliar_matriz(caminho, registros):
    matriz = exigir_fora_de_git(caminho, "Matriz Analítica 2025")
    if not matriz.is_file() or matriz.suffix.lower() != ".xlsx":
        raise ErroPipeline("Matriz Analítica deve ser um arquivo xlsx")
    try:
        import openpyxl
    except ImportError as exc:
        raise ErroPipeline("openpyxl não instalado") from exc

    esperados = agregados_matriz_esperados(registros)
    controle = {"campos_comparados": 0, "divergencias": []}
    wb = openpyxl.load_workbook(matriz, data_only=True, read_only=True)
    try:
        obrigatorias = {"Resumo_Executivo", "Mensal", "Rede_por_Terapia"}
        faltantes = obrigatorias - set(wb.sheetnames)
        if faltantes:
            raise ErroPipeline(
                f"Matriz sem abas obrigatórias: {sorted(faltantes)}"
            )

        # Resumo_Executivo: 14 campos por terapia.
        ws = wb["Resumo_Executivo"]
        linha, cab = localizar_cabecalho_matriz(ws, "Terapia")
        campos_resumo = [
            "Registros", "Beneficiários*", "Sessões", "Pagamento",
            "Ben. própria", "Sessões própria", "Média própria",
            "Mediana própria", "Ben. externa", "Sessões externa",
            "Média externa", "Mediana externa", "Custo/sessão própria",
            "Custo/sessão externa",
        ]
        indices = {
            campo: indice_cabecalho(cab, campo, ws.title)
            for campo in campos_resumo
        }
        idx_codigo = indice_cabecalho(cab, "Código", ws.title)
        vistos = set()
        for row in ws.iter_rows(min_row=linha + 1, values_only=True):
            terapia_id = id_terapia_por_codigo_matriz(
                row[idx_codigo] if idx_codigo < len(row) else None
            )
            if not terapia_id:
                continue
            if terapia_id in vistos:
                raise ErroPipeline(
                    f"Matriz, Resumo_Executivo: terapia duplicada {terapia_id}"
                )
            vistos.add(terapia_id)
            esperado = esperados["Resumo_Executivo"].get(terapia_id)
            if esperado is None:
                raise ErroPipeline(
                    f"Matriz contém terapia sem agregado 2025: {terapia_id}"
                )
            for campo, idx in indices.items():
                observado = decimal_matriz(row[idx] if idx < len(row) else None)
                comparar_celula(
                    controle, ws.title, terapia_id, campo,
                    observado, esperado[campo],
                )
        if vistos != set(esperados["Resumo_Executivo"]):
            raise ErroPipeline(
                "Matriz, Resumo_Executivo: conjunto de terapias divergente"
            )

        # Mensal: oito campos por terapia e competência.
        ws = wb["Mensal"]
        linha, cab = localizar_cabecalho_matriz(ws, "Terapia")
        campos_mensal = [
            "Registros", "Beneficiários", "Sessões", "Pagamento",
            "Ben. própria", "Sessões própria", "Ben. externa",
            "Sessões externa",
        ]
        indices = {
            campo: indice_cabecalho(cab, campo, ws.title)
            for campo in campos_mensal
        }
        idx_terapia = indice_cabecalho(cab, "Terapia", ws.title)
        idx_mes = indice_cabecalho(cab, "Competência", ws.title)
        vistos = set()
        for row in ws.iter_rows(min_row=linha + 1, values_only=True):
            terapia_id = id_terapia_por_nome_matriz(
                row[idx_terapia] if idx_terapia < len(row) else None
            )
            if not terapia_id:
                continue
            mes = texto_celula(row[idx_mes] if idx_mes < len(row) else None)
            if not mes.startswith("2025/"):
                continue
            chave = (terapia_id, mes)
            if chave in vistos:
                raise ErroPipeline(f"Matriz, Mensal: linha duplicada {chave}")
            vistos.add(chave)
            esperado = esperados["Mensal"].get(chave)
            if esperado is None:
                raise ErroPipeline(
                    f"Matriz, Mensal: célula sem correspondente no pipeline: {chave}"
                )
            for campo, idx in indices.items():
                observado = decimal_matriz(row[idx] if idx < len(row) else None)
                comparar_celula(
                    controle, ws.title, chave, campo,
                    observado, esperado[campo],
                )
        if vistos != set(esperados["Mensal"]):
            raise ErroPipeline("Matriz, Mensal: conjunto de linhas divergente")

        # Rede_por_Terapia: 13 campos por linha da referência. Esta leitura
        # existe apenas para reconciliação; o JSON do produto mantém rede única.
        ws = wb["Rede_por_Terapia"]
        linha, cab = localizar_cabecalho_matriz(ws, "Terapia")
        campos_rede = [
            "Beneficiários", "Sessões", "Pagamento", "Média", "Mediana",
            "P25", "P75", "Mínimo", "Máximo", "Desvio-padrão",
            "Custo/sessão", "Custo/beneficiário", "Sobreposição",
        ]
        indices = {
            campo: indice_cabecalho(cab, campo, ws.title)
            for campo in campos_rede
        }
        idx_terapia = indice_cabecalho(cab, "Terapia", ws.title)
        idx_rede = indice_cabecalho(cab, "Rede", ws.title)
        vistos = set()
        for row in ws.iter_rows(min_row=linha + 1, values_only=True):
            terapia_id = id_terapia_por_nome_matriz(
                row[idx_terapia] if idx_terapia < len(row) else None
            )
            if not terapia_id:
                continue
            nome_rede = texto_celula(row[idx_rede] if idx_rede < len(row) else None)
            chave = (terapia_id, nome_rede)
            if chave in vistos:
                raise ErroPipeline(
                    f"Matriz, Rede_por_Terapia: linha duplicada {chave}"
                )
            vistos.add(chave)
            esperado = esperados["Rede_por_Terapia"].get(chave)
            if esperado is None:
                raise ErroPipeline(
                    f"Matriz, Rede_por_Terapia: linha sem correspondente: {chave}"
                )
            for campo, idx in indices.items():
                observado = decimal_matriz(row[idx] if idx < len(row) else None)
                comparar_celula(
                    controle, ws.title, chave, campo,
                    observado, esperado[campo],
                )
        if vistos != set(esperados["Rede_por_Terapia"]):
            raise ErroPipeline(
                "Matriz, Rede_por_Terapia: conjunto de linhas divergente"
            )
    finally:
        wb.close()

    if controle["divergencias"]:
        amostra = "\n  - ".join(controle["divergencias"][:12])
        raise ErroPipeline(
            f"Matriz 2025 divergiu em {len(controle['divergencias'])} campos:\n"
            f"  - {amostra}"
        )
    return {
        "arquivo": matriz.name,
        "sha256": sha256_arquivo(matriz),
        "abas": [
            "Resumo_Executivo", "Mensal", "Rede_por_Terapia"
        ],
        "campos_comparados": controle["campos_comparados"],
        "status": "reconciliada",
        "escopo": (
            "2025, 18 terapias de origem; Método PECS — Fonoaudiologia "
            "validada na Matriz e descartada do painel por cobertura incompleta"
        ),
    }


# ---------------- validações bloqueantes ----------------


def primeiro_padrao_encontrado(texto, valores):
    """Busca linear Aho-Corasick sem devolver o dado sensível encontrado."""
    # A saída usa dNome em vários rótulos. Normalizar os dois lados impede que
    # a retirada de acentos transforme um identificador em falso negativo.
    unicos = {d_nome(v) for v in valores if len(v.strip()) >= 6}
    if not unicos:
        return False
    proximos = [{}]
    falhas = [0]
    finais = [False]
    for valor in unicos:
        estado = 0
        for caractere in valor:
            if caractere not in proximos[estado]:
                proximos[estado][caractere] = len(proximos)
                proximos.append({})
                falhas.append(0)
                finais.append(False)
            estado = proximos[estado][caractere]
        finais[estado] = True
    fila = deque()
    for estado in proximos[0].values():
        fila.append(estado)
    while fila:
        atual = fila.popleft()
        for caractere, destino in proximos[atual].items():
            fila.append(destino)
            retorno = falhas[atual]
            while retorno and caractere not in proximos[retorno]:
                retorno = falhas[retorno]
            falhas[destino] = proximos[retorno].get(caractere, 0)
            finais[destino] = finais[destino] or finais[falhas[destino]]
    estado = 0
    for caractere in d_nome(texto):
        while estado and caractere not in proximos[estado]:
            estado = falhas[estado]
        estado = proximos[estado].get(caractere, 0)
        if finais[estado]:
            return True
    return False


def validar_anti_vazamento(texto, registros):
    campos = {
        "nome de beneficiário": {r["nome"] for r in registros},
        "matrícula": {r["matricula"] for r in registros},
        "nascimento": {r["nasc"] for r in registros},
        "guia": {r["guia"] for r in registros},
    }
    for rotulo, valores in campos.items():
        if primeiro_padrao_encontrado(texto, valores):
            raise ErroPipeline(
                f"anti-vazamento bloqueou presença de {rotulo} no artefato"
            )


def validar_estrutura_sem_identificadores(dados):
    proibidas = {
        "nome_beneficiario", "nome_beneficiário", "matricula", "matrícula",
        "nascimento", "data_nascimento", "guia", "numero_guia",
        "número_guia", "paciente", "chave_paciente",
    }

    def percorrer(no, caminho="dados"):
        if isinstance(no, dict):
            for chave, valor in no.items():
                if chave.casefold() in proibidas:
                    raise ErroPipeline(
                        f"campo identificador proibido no artefato: {caminho}.{chave}"
                    )
                percorrer(valor, f"{caminho}.{chave}")
        elif isinstance(no, list):
            for indice, valor in enumerate(no):
                percorrer(valor, f"{caminho}[{indice}]")

    percorrer(dados)


def validar_saida(dados, auditoria, reconciliacao, exigir_referencias):
    erros = []
    meta = dados["meta"]
    if len(meta["terapias"]) != 17:
        erros.append("JSON não contém exatamente 17 terapias")
    if len(meta["manifesto_fontes"]) != 19:
        erros.append("manifesto não contém exatamente 19 fontes")
    if len(meta["periodos"]) != 30 or len(dados["recortes"]) != 30:
        erros.append("contrato de períodos deve conter 30 recortes")
    descartadas = meta["terapias_descartadas"]
    if len(descartadas) != 1 or descartadas[0]["id"] != ID_PECS:
        erros.append("descarte programático de PECS não está único")

    # Contrato v4: a soma das fatias prestador×disciplina reproduz o total do
    # prestador, e a soma das séries por disciplina reproduz a série mensal.
    for periodo_id, recorte in dados["recortes"].items():
        fatias = recorte.get("prestador_disciplina")
        if not isinstance(fatias, list):
            erros.append(f"{periodo_id}: prestador_disciplina ausente")
            continue
        soma_prest = defaultdict(lambda: [Decimal(0), Decimal(0)])
        for fatia in fatias:
            soma_prest[fatia["nome"]][0] += Decimal(str(fatia["pagamento"]))
            soma_prest[fatia["nome"]][1] += Decimal(str(fatia["sessoes"]))
        for prestador in recorte["prestadores"]:
            pago, sess = soma_prest.get(prestador["nome"], [Decimal(0), Decimal(0)])
            if abs(pago - Decimal(str(prestador["pagamento"]))) > Decimal("0.02"):
                erros.append(
                    f"{periodo_id}: fatias de {prestador['nome']} não somam o "
                    f"pagamento total ({pago} != {prestador['pagamento']})"
                )
            if sess != Decimal(str(prestador["sessoes"])):
                erros.append(
                    f"{periodo_id}: fatias de {prestador['nome']} não somam as sessões"
                )
        series_disc = recorte.get("disciplina_serie")
        if not isinstance(series_disc, dict) or not series_disc:
            erros.append(f"{periodo_id}: disciplina_serie ausente")
        else:
            for indice, ponto in enumerate(recorte["serie_mensal"]):
                soma_pag = sum(
                    (Decimal(str(serie[indice]["pagamento"] or 0)) for serie in series_disc.values()),
                    Decimal(0),
                )
                if abs(soma_pag - Decimal(str(ponto["pagamento"] or 0))) > Decimal("0.02"):
                    erros.append(
                        f"{periodo_id}: disciplina_serie não soma a série mensal em {ponto['m']}"
                    )
                    break
        if not isinstance(recorte.get("referencias"), dict):
            erros.append(f"{periodo_id}: referencias por disciplina ausentes")

        # Contrato v5: consolidação da Casa Unimed coerente com a composição.
        resumo = recorte["resumo"]
        casa = next(
            (x for x in recorte["prestadores"] if x["nome"] == CASA_UNIMED),
            None,
        )
        if casa is not None:
            composicao = casa.get("composicao") or []
            soma_pag = sum(
                (Decimal(str(p["pagamento"])) for p in composicao), Decimal(0)
            )
            soma_sess = sum(
                (Decimal(str(p["sessoes"])) for p in composicao), Decimal(0)
            )
            if abs(soma_pag - Decimal(str(casa["pagamento"]))) > Decimal("0.02"):
                erros.append(
                    f"{periodo_id}: composição da Casa Unimed não soma o "
                    f"pagamento da entidade ({soma_pag} != {casa['pagamento']})"
                )
            if soma_sess != Decimal(str(casa["sessoes"])):
                erros.append(
                    f"{periodo_id}: composição da Casa Unimed não soma as sessões"
                )
            if casa["canais"] != [CANAL_CASA_UNIMED]:
                erros.append(
                    f"{periodo_id}: entidade Casa Unimed com canal fora de "
                    f"{CANAL_CASA_UNIMED}"
                )
            if len(composicao) != casa["quantidade_entidades"]:
                erros.append(
                    f"{periodo_id}: quantidade_entidades diverge da composição"
                )
            if len(composicao) != resumo["profissionais_casa_unimed"]:
                erros.append(
                    f"{periodo_id}: profissionais_casa_unimed diverge da composição"
                )
            serie_casa = recorte["serie_prestador"].get(CASA_UNIMED)
            soma_serie = sum(
                (Decimal(str(p["pagamento"] or 0)) for p in serie_casa or []),
                Decimal(0),
            )
            if serie_casa is None or abs(
                soma_serie - Decimal(str(casa["pagamento"]))
            ) > Decimal("0.02"):
                erros.append(
                    f"{periodo_id}: serie_prestador da Casa Unimed não soma o pagamento"
                )
        elif resumo["profissionais_casa_unimed"]:
            erros.append(
                f"{periodo_id}: profissionais da Casa contados sem entidade publicada"
            )
        for item in recorte["prestadores"]:
            if item["nome"] != CASA_UNIMED and CANAL_CASA_UNIMED in item["canais"]:
                erros.append(
                    f"{periodo_id}: registro do canal Casa Unimed escapou da "
                    f"consolidação em {item['nome']}"
                )
        audit_recorte = auditoria["recortes"][periodo_id]
        esperado_publicados = (
            resumo["prestadores_consolidados"]
            - audit_recorte["profissionais_exclusivos_casa"]
            + (1 if casa is not None else 0)
        )
        if resumo["prestadores_publicados"] != esperado_publicados:
            erros.append(
                f"{periodo_id}: prestadores_publicados fora do invariante "
                "consolidados - exclusivos da Casa + 1"
            )
        if resumo["prestadores_publicados"] != len(recorte["prestadores"]):
            erros.append(
                f"{periodo_id}: prestadores_publicados diverge do ranking"
            )

        # Contrato v5: cross-foot de disciplina_canais e resumo_disciplina.
        disc_canais = recorte.get("disciplina_canais")
        resumo_disc = recorte.get("resumo_disciplina")
        if not isinstance(series_disc, dict):
            series_disc = {}
        if not isinstance(disc_canais, dict) or not isinstance(resumo_disc, dict):
            erros.append(
                f"{periodo_id}: disciplina_canais/resumo_disciplina ausentes"
            )
        else:
            if not (set(disc_canais) == set(resumo_disc) == set(series_disc)):
                erros.append(
                    f"{periodo_id}: conjuntos de disciplinas divergentes entre "
                    "disciplina_canais, resumo_disciplina e disciplina_serie"
                )
            soma_fatias = defaultdict(lambda: [Decimal(0), Decimal(0)])
            for fatia in fatias:
                soma_fatias[fatia["disciplina"]][0] += Decimal(str(fatia["pagamento"]))
                soma_fatias[fatia["disciplina"]][1] += Decimal(str(fatia["sessoes"]))
            soma_canal = defaultdict(lambda: [Decimal(0), Decimal(0)])
            for disciplina, itens_canal in disc_canais.items():
                pago = sum(
                    (Decimal(str(x["pagamento"] or 0)) for x in itens_canal),
                    Decimal(0),
                )
                sess = sum(
                    (Decimal(str(x["sessoes"] or 0)) for x in itens_canal),
                    Decimal(0),
                )
                pago_fatias, sess_fatias = soma_fatias.get(
                    disciplina, [Decimal(0), Decimal(0)]
                )
                if abs(pago - pago_fatias) > Decimal("0.02") or sess != sess_fatias:
                    erros.append(
                        f"{periodo_id}: disciplina_canais de {disciplina} não "
                        "soma as fatias prestador x disciplina"
                    )
                for x in itens_canal:
                    soma_canal[x["canal"]][0] += Decimal(str(x["pagamento"] or 0))
                    soma_canal[x["canal"]][1] += Decimal(str(x["sessoes"] or 0))
            for item in recorte["canais"]:
                pago, sess = soma_canal.get(
                    item["canal"], [Decimal(0), Decimal(0)]
                )
                if abs(pago - Decimal(str(item["pagamento"] or 0))) > Decimal(
                    "0.02"
                ) or sess != Decimal(str(item["sessoes"] or 0)):
                    erros.append(
                        f"{periodo_id}: disciplina_canais não cruza com o "
                        f"canal {item['canal']}"
                    )
            for disciplina, item in resumo_disc.items():
                serie = series_disc.get(disciplina) or []
                pago = sum(
                    (Decimal(str(p["pagamento"] or 0)) for p in serie),
                    Decimal(0),
                )
                sess = sum(
                    (Decimal(str(p["sessoes"] or 0)) for p in serie), Decimal(0)
                )
                if abs(pago - Decimal(str(item["pagamento"] or 0))) > Decimal(
                    "0.02"
                ) or sess != Decimal(str(item["sessoes"] or 0)):
                    erros.append(
                        f"{periodo_id}: resumo_disciplina de {disciplina} não "
                        "soma a série da disciplina"
                    )
        pacientes_disc = recorte.get("pacientes_disciplina")
        if not isinstance(pacientes_disc, dict) or (
            isinstance(resumo_disc, dict) and set(pacientes_disc) != set(resumo_disc)
        ):
            erros.append(f"{periodo_id}: pacientes_disciplina incompleto")

        # Contrato v5: kits publicados somente com k>=5 e somas coerentes.
        kits = recorte.get("kits")
        if not isinstance(kits, dict):
            erros.append(f"{periodo_id}: kits ausentes")
        else:
            ids_catalogo = {k["id"] for k in KITS}
            total_criancas = 0
            for kit_id, item in kits.items():
                if kit_id not in ids_catalogo or item["id"] != kit_id:
                    erros.append(f"{periodo_id}: kit fora do catálogo: {kit_id}")
                    continue
                if item["criancas"] < K_MINIMO_KIT:
                    erros.append(
                        f"{periodo_id}: kit {kit_id} publicado com menos de "
                        f"{K_MINIMO_KIT} crianças"
                    )
                if item["criancas"] != item["pacientes"]:
                    erros.append(
                        f"{periodo_id}: kit {kit_id} com criancas != pacientes"
                    )
                total_criancas += item["criancas"]
                pago = sum(
                    (Decimal(str(x["pagamento"] or 0)) for x in item["canais"]),
                    Decimal(0),
                )
                sess = sum(
                    (Decimal(str(x["sessoes"] or 0)) for x in item["canais"]),
                    Decimal(0),
                )
                if abs(pago - Decimal(str(item["pagamento"] or 0))) > Decimal(
                    "0.02"
                ) or sess != Decimal(str(item["sessoes"] or 0)):
                    erros.append(
                        f"{periodo_id}: canais do kit {kit_id} não somam o total"
                    )
                recomputado = razao(
                    Decimal(str(item["pagamento"] or 0)), item["crianca_meses"]
                )
                if recomputado is None or abs(
                    Decimal(str(recomputado))
                    - Decimal(str(item["custo_crianca_mes"]))
                ) > Decimal("0.01"):
                    erros.append(
                        f"{periodo_id}: custo_crianca_mes do kit {kit_id} não "
                        "reproduz pagamento/criança-mês"
                    )
            if total_criancas > resumo["pacientes"]:
                erros.append(
                    f"{periodo_id}: kits somam mais crianças que o recorte"
                )
    if reconciliacao.get("status") != "reconciliada":
        erros.append("Matriz 2025 não reconciliada")
    if "projecao" in json.dumps(dados, ensure_ascii=False).lower():
        erros.append("projeções não podem compor o produto")
    for recorte in dados["recortes"].values():
        for item in recorte["prestadores"]:
            if d_nome(item["nome"]) != item["nome"]:
                erros.append("rótulo de prestador fora da regra dNome")
                break
        for item in recorte["solicitantes"]["itens"]:
            if d_nome(item["nome"]) != item["nome"]:
                erros.append("rótulo de solicitante fora da regra dNome")
                break

    global_audit = auditoria["recortes"]["tudo"]
    metricas = global_audit["metricas"]
    linhas = global_audit["linhas_cuidado"]
    if exigir_referencias:
        referencias = [
            (
                metricas["n_pacientes"] == 1416,
                f"crianças únicas: {metricas['n_pacientes']} != 1416",
            ),
            (
                global_audit["prestadores_observados"] == 161,
                "prestadores observados devem ser 161",
            ),
            (
                global_audit["prestadores_consolidados"] == 160,
                "prestadores consolidados devem ser 160",
            ),
            # Contrato v5, fixado no primeiro build real após conferência
            # manual: 148 = 160 consolidados - 13 exclusivos da Casa + 1.
            (
                global_audit["prestadores_publicados"] == 148,
                "entidades publicadas devem ser 148",
            ),
            (
                dados["recortes"]["tudo"]["resumo"]["profissionais_casa_unimed"]
                == 14,
                "a Casa Unimed deve consolidar 14 profissionais",
            ),
            (
                metricas["sessoes"] == Decimal("81217"),
                f"sessões: {metricas['sessoes']} != 81217",
            ),
            (
                metricas["pagamento"].quantize(CENTAVO) == Decimal("13919882.74"),
                "pagamento total diverge de R$ 13.919.882,74",
            ),
            (
                linhas["distribuicao_clinica"]
                == {1: 730, 2: 330, 3: 248, 4: 71, 5: 22, 6: 13, 7: 2},
                "distribuição clínica de linhas de cuidado diverge",
            ),
            (
                linhas["exposicoes_clinicas"] == 2620,
                "exposições clínicas devem ser 2.620",
            ),
            (
                linhas["distribuicao_componentes"]
                == {1: 716, 2: 332, 3: 246, 4: 81, 5: 26, 6: 13, 7: 2},
                "reconciliação técnica por códigos diverge",
            ),
            (
                linhas["exposicoes_componentes"] == 2664,
                "exposições técnicas devem ser 2.664",
            ),
            (
                reconciliacao.get("campos_comparados") == 2247,
                "Matriz deve reconciliar exatamente 2.247 campos",
            ),
        ]
        erros.extend(mensagem for ok, mensagem in referencias if not ok)
    if erros:
        raise ErroPipeline("validações reprovadas:\n  - " + "\n  - ".join(erros))

    validar_estrutura_sem_identificadores(dados)
    texto = json.dumps(dados, ensure_ascii=False, sort_keys=True)
    if len(texto.encode("utf-8")) > JSON_MAX_BYTES:
        raise ErroPipeline(
            f"JSON excede o limite de {JSON_MAX_BYTES // (1024 * 1024)} MB"
        )
    validar_anti_vazamento(texto, auditoria["registros_todos"])
    return texto


# ---------------- saídas atômicas ----------------


def escrita_atomica(caminho, conteudo):
    destino = Path(caminho)
    destino.parent.mkdir(parents=True, exist_ok=True)
    descritor, temporario = tempfile.mkstemp(
        prefix=f".{destino.name}.", suffix=".tmp", dir=destino.parent
    )
    try:
        with os.fdopen(descritor, "w", encoding="utf-8", newline="\n") as f:
            f.write(conteudo)
            f.flush()
            os.fsync(f.fileno())
        os.replace(temporario, destino)
    finally:
        if os.path.exists(temporario):
            os.unlink(temporario)


def escrita_atomica_multipla(saidas):
    """Troca um conjunto de textos com rollback se qualquer troca falhar."""
    originais = {}
    for caminho, _ in saidas:
        destino = Path(caminho)
        originais[destino] = (
            destino.read_text(encoding="utf-8") if destino.exists() else None
        )

    alterados = []
    try:
        for caminho, conteudo in saidas:
            destino = Path(caminho)
            escrita_atomica(destino, conteudo)
            alterados.append(destino)
    except Exception as exc:
        falhas_rollback = []
        for destino in reversed(alterados):
            try:
                original = originais[destino]
                if original is None:
                    destino.unlink(missing_ok=True)
                else:
                    escrita_atomica(destino, original)
            except Exception as rollback_exc:  # pragma: no cover - falha de SO
                falhas_rollback.append(f"{destino}: {rollback_exc}")
        if falhas_rollback:
            raise ErroPipeline(
                "falha de escrita e rollback incompleto: "
                + "; ".join(falhas_rollback)
            ) from exc
        raise


# Gate visual preservado. O broker privado valida as mesmas credenciais no
# csv-open-auth e emite uma sessão opaca, restrita ao painel, por quatro horas.
GATE_P_HTML = """<div id="gate" role="dialog" aria-modal="true" aria-labelledby="gate-title" style="position:fixed;inset:0;z-index:99999;background:#003c3a;display:flex;align-items:center;justify-content:center;padding:20px;overflow:auto;font-family:'Source Sans 3','Segoe UI',Arial,sans-serif">
<form id="gate-form" style="background:#fff;border-radius:16px;max-width:400px;width:100%;padding:36px 32px;box-shadow:0 24px 64px rgba(0,0,0,.4);margin:auto">
<img src="https://assets.grupocsv.com/logos/unimed-gv/sem-box-pinheiro.png" alt="Unimed GV" style="height:44px;margin-bottom:18px">
<h1 id="gate-title" style="font-size:19px;color:#004e4c;margin:0 0 6px">Painel Terapias Especiais — acesso restrito</h1>
<p style="font-size:13px;color:#5c7f7d;margin:0 0 20px">Uso exclusivo Unimed Governador Valadares. Informe seu e-mail corporativo e a senha de acesso.</p>
<label for="gate-email" style="display:block;font-size:11px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;color:#5c7f7d;margin-bottom:5px">E-mail</label>
<input id="gate-email" type="email" required placeholder="nome@unimedgv.com.br" autocomplete="email" style="width:100%;box-sizing:border-box;font-size:16px;padding:10px 12px;border:1px solid #dfe3db;border-radius:10px;margin-bottom:14px">
<label for="gate-pass" style="display:block;font-size:11px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;color:#5c7f7d;margin-bottom:5px">Senha</label>
<div style="position:relative;margin-bottom:18px">
<input id="gate-pass" type="password" required autocomplete="current-password" style="width:100%;box-sizing:border-box;font-size:16px;padding:10px 52px 10px 12px;border:1px solid #dfe3db;border-radius:10px">
<button type="button" id="gate-eye" aria-label="Mostrar senha" aria-pressed="false" style="position:absolute;right:0;top:50%;transform:translateY(-50%);width:44px;height:44px;background:none;border:0;cursor:pointer;padding:12px;color:#5c7f7d;line-height:0">
<svg id="gate-eye-on" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"/><circle cx="12" cy="12" r="3"/></svg>
<svg id="gate-eye-off" width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="display:none"><path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94"/><path d="M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19"/><line x1="1" y1="1" x2="23" y2="23"/></svg>
</button>
</div>
<button type="submit" id="gate-btn" style="width:100%;background:#00995d;color:#fff;border:0;border-radius:9999px;padding:12px;font-size:14px;font-weight:700;cursor:pointer">Entrar</button>
<p id="gate-msg" role="status" aria-live="polite" style="font-size:12.5px;color:#c0392b;min-height:16px;margin:12px 0 0"></p>
</form></div>
<script>(function(){'use strict';var PAGE='painel-tea',KEY='oa_sess_'+PAGE,AUTH='https://unimed-te-data.guilherme-thom.workers.dev/v1/session/open';
function ok(){var g=document.getElementById('gate');if(g)g.remove();document.body.style.overflow='';}
function ler(){try{return localStorage.getItem(KEY)||sessionStorage.getItem(KEY)||'';}catch(e){try{return sessionStorage.getItem(KEY)||'';}catch(_e){return '';}}}
function salvar(value){try{localStorage.setItem(KEY,value);return true;}catch(e){try{sessionStorage.setItem(KEY,value);return true;}catch(_e){return false;}}}
try{var s=JSON.parse(ler()||'null');if(s&&s.token&&s.exp>Date.now()){ok();return;}}catch(e){}
document.body.style.overflow='hidden';
var eye=document.getElementById('gate-eye');
eye.addEventListener('click',function(){var inp=document.getElementById('gate-pass'),m=inp.type==='password';inp.type=m?'text':'password';eye.setAttribute('aria-pressed',String(m));eye.setAttribute('aria-label',m?'Ocultar senha':'Mostrar senha');document.getElementById('gate-eye-on').style.display=m?'none':'';document.getElementById('gate-eye-off').style.display=m?'':'none';inp.focus();});
document.getElementById('gate-form').addEventListener('submit',function(ev){ev.preventDefault();
var btn=document.getElementById('gate-btn'),msg=document.getElementById('gate-msg');btn.disabled=true;btn.textContent='Verificando\\u2026';msg.textContent='';
fetch(AUTH,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({page:PAGE,email:document.getElementById('gate-email').value.trim(),senha:document.getElementById('gate-pass').value}),signal:AbortSignal.timeout(20000)})
.then(function(r){return r.json().catch(function(){return{};});})
.then(function(j){if(j&&j.ok&&j.token){var exp=Date.parse(j.expires_at);if(!Number.isFinite(exp))exp=Date.now()+4*3600*1000;if(!salvar(JSON.stringify({token:j.token,exp:exp}))){msg.textContent='O navegador bloqueou o armazenamento da sessão. Ajuste as permissões e tente novamente.';btn.disabled=false;btn.textContent='Entrar';return;}document.getElementById('gate-pass').value='';ok();}
else{msg.textContent=(j&&(j.msg||j.error))||'Acesso negado. Confira e-mail e senha.';btn.disabled=false;btn.textContent='Entrar';}})
.catch(function(){msg.textContent='Falha de rede. Tente novamente.';btn.disabled=false;btn.textContent='Entrar';});});document.getElementById('gate-email').focus();})();</script>"""


def emitir_p():
    if not HTML_ALVO.is_file():
        raise ErroPipeline(f"HTML interno não encontrado: {HTML_ALVO}")
    html = HTML_ALVO.read_text(encoding="utf-8")
    if re.search(r'id=["\']tea-data["\']', html, flags=re.I):
        raise ErroPipeline(
            "HTML interno ainda contém tea-data; remova o payload antes de gerar /p/"
        )
    html = re.sub(r'<span id="hub-auth-slot"[^>]*></span>\s*', "", html)
    html = re.sub(
        r'<script[^>]*src="[^"]*hub-auth\.js"[^>]*></script>\s*', "", html
    )
    html = re.sub(
        r'<a [^>]*class="[^"]*voltar-hub[^"]*"[^>]*>.*?</a>\s*',
        "",
        html,
        flags=re.S,
    )
    html = html.replace(
        'content="https://hub.grupocsv.com/unimed/tea.html"',
        'content="https://hub.grupocsv.com/p/painel-tea/"',
    )
    if 'id="gate"' not in html:
        html = html.replace("</body>", GATE_P_HTML + "\n</body>")
    if '<meta name="robots" content="noindex, nofollow">' not in html:
        raise ErroPipeline("cópia /p/ perderia meta robots noindex, nofollow")
    if 'id="gate"' not in html:
        raise ErroPipeline("gate embutido não foi preservado na cópia /p/")
    if re.search(r'<script[^>]*hub-auth\.js', html):
        raise ErroPipeline("cópia /p/ não pode carregar hub-auth.js")
    contratos_gate = {
        "slug": "var PAGE='painel-tea'",
        "sessão de 4 horas": "4*3600*1000",
        "e-mail corporativo": "@unimedgv.com.br",
        "controle de senha": 'id="gate-eye"',
        "endpoint de autenticação": "unimed-te-data.guilherme-thom.workers.dev/v1/session/open",
        "token opaco": "JSON.stringify({token:j.token,exp:exp})",
    }
    ausentes = [rotulo for rotulo, trecho in contratos_gate.items() if trecho not in html]
    if ausentes:
        raise ErroPipeline(
            "cópia /p/ perdeu contrato do gate: " + ", ".join(ausentes)
        )
    destino = P_DIR / "index.html"
    escrita_atomica(destino, html)
    print(f"cópia /p/: {destino} ({len(html.encode('utf-8')) // 1024} KB)")
    return destino


def formatar_inteiro(valor):
    return f"{int(valor):,}".replace(",", ".")


def formatar_brl(valor):
    quantizado = Decimal(valor).quantize(CENTAVO, rounding=ROUND_HALF_UP)
    base = f"{quantizado:,.2f}"
    base = base.replace(",", "#").replace(".", ",").replace("#", ".")
    return f"R$ {base}"


def formatar_pct(valor):
    return f"{Decimal(str(valor)).quantize(Decimal('0.1'))}%".replace(".", ",")


def formatar_distribuicao(distribuicao):
    return " / ".join(
        f"{quantidade}={formatar_inteiro(distribuicao.get(quantidade, 0))}"
        for quantidade in sorted(distribuicao)
    )


def imprimir_relatorio(dados, auditoria, reconciliacao, saida):
    global_audit = auditoria["recortes"]["tudo"]
    m = global_audit["metricas"]
    l = global_audit["linhas_cuidado"]
    print("=" * 78)
    print("RELATÓRIO DO BUILD — PAINEL TERAPIAS ESPECIAIS")
    print(f"Período canônico       : {INICIO_CANONICO} a {FIM_CANONICO}")
    print(f"Fontes                 : {len(dados['meta']['manifesto_fontes'])}")
    print(f"Terapias mantidas      : {len(dados['meta']['terapias'])}")
    descarte = dados["meta"]["terapias_descartadas"][0]
    print(
        "Terapia descartada     : "
        f"{descarte['terapia']} — {descarte['motivo']}"
    )
    print(f"Crianças únicas        : {formatar_inteiro(m['n_pacientes'])}")
    print(
        "Prestadores            : "
        f"{formatar_inteiro(global_audit['prestadores_observados'])} "
        "nomes-fonte / "
        f"{formatar_inteiro(global_audit['prestadores_consolidados'])} "
        "entidades após consolidar o par de CNPJ"
    )
    recorte_tudo = dados["recortes"]["tudo"]
    resumo_tudo = recorte_tudo["resumo"]
    print(
        "Entidades publicadas   : "
        f"{formatar_inteiro(resumo_tudo['prestadores_publicados'])} "
        f"(Casa Unimed consolida "
        f"{formatar_inteiro(resumo_tudo['profissionais_casa_unimed'])} "
        "profissionais)"
    )
    casa = next(
        (x for x in recorte_tudo["prestadores"] if x["nome"] == CASA_UNIMED),
        None,
    )
    if casa is not None:
        print(
            "Casa Unimed            : "
            f"{formatar_brl(casa['pagamento'])} / "
            f"{formatar_inteiro(casa['sessoes'])} sessões na entidade"
        )
    print("Disciplina x canal     : ok (cross-foot canais/disciplinas)")
    kits = recorte_tudo["kits"]
    print(
        "Kits publicados        : "
        + (
            "; ".join(
                f"{item['id']}={formatar_inteiro(item['criancas'])} crianças"
                for item in kits.values()
            )
            or "nenhum"
        )
        + f" | suprimidos por k<{K_MINIMO_KIT}: {len(KITS) - len(kits)}"
    )
    print(f"Sessões                : {formatar_inteiro(m['sessoes'])}")
    print(f"Pagamento              : {formatar_brl(m['pagamento'])}")
    print(f"Sessões no denominador : {formatar_inteiro(m['sessoes_pagas'])}")
    print(
        "Linhas clínicas (17)   : "
        f"{formatar_distribuicao(l['distribuicao_clinica'])} "
        f"| exposições={formatar_inteiro(l['exposicoes_clinicas'])} "
        f"| 2+={formatar_pct(pct(l['dois_mais_clinico'], m['n_pacientes']))}"
    )
    print(
        "Componentes por código : "
        f"{formatar_distribuicao(l['distribuicao_componentes'])} "
        f"| exposições={formatar_inteiro(l['exposicoes_componentes'])} "
        f"| 2+={formatar_pct(pct(l['dois_mais_componentes'], m['n_pacientes']))}"
    )
    print(
        "Nota de reconciliação  : componentes por código são um rastro técnico; "
        "AT integra ABA — Psicologia e não é terapia clínica adicional"
    )
    print(
        "Matriz Analítica 2025  : "
        f"{reconciliacao['status']} em "
        f"{formatar_inteiro(reconciliacao['campos_comparados'])} campos"
    )
    print("Hashes SHA-256:")
    for item in dados["meta"]["manifesto_fontes"]:
        print(f"  {item['sha256']}  {item['arquivo']}")
    print(f"  {reconciliacao['sha256']}  {reconciliacao['arquivo']}")
    print(f"JSON privado           : {saida}")
    print("HTML/repositório       : nenhum dado gravado")
    print("Validações             : todas aprovadas")
    print("=" * 78)


# ---------------- fixture sintética multi-fonte ----------------


def exigir_teste(condicao, mensagem):
    if not condicao:
        raise ErroPipeline(f"selftest: {mensagem}")


def valor_excel(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


def criar_fontes_sinteticas(diretorio):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ErroPipeline("openpyxl não instalado") from exc

    diretorio.mkdir(parents=True, exist_ok=True)
    # Pacientes 1–6 frequentam todas as fontes (kit tripé ampliado); 7–11
    # existem só nas três fontes das disciplinas principais (kit tripé, k==5);
    # 12–13 existem só em Psicologia (kit "unica", k<5 — não publica).
    pacientes = [
        {
            "nome": f"CRIANCA SINTETICA {i + 1}",
            "nasc": date(1990, 1, 1) if i == 5 else date(2014, i % 12 + 1, 1),
            "genero": "FEMININO" if i % 2 else "MASCULINO",
        }
        for i in range(13)
    ]
    fontes_tripe = {"50005103", "50005189", "50005170"}
    for indice_codigo, codigo in enumerate(MAPA_CODIGOS):
        wb = Workbook()
        ws = wb.active
        if indice_codigo == 0:
            ws.title = "LEIA-ME"
            ws["A1"] = "Instruções sintéticas"
            ws = wb.create_sheet("Dados")
            ws.append(["Fixture sintética; cabeçalho abaixo"])
            ws.append([])
        ws.append(CABECALHO_FONTE)
        meses = (
            sorted(MESES_PECS_ESPERADOS)
            if codigo == "50005146"
            else list(MESES_CANONICOS)
        )
        for mes in meses:
            ano, numero_mes = map(int, mes.split("/"))
            for i, paciente in enumerate(pacientes):
                if 6 <= i <= 10 and codigo not in fontes_tripe:
                    continue
                if i >= 11 and codigo != "50005103":
                    continue
                if codigo == COD_AT:
                    prestador = PEDIAKIDS
                elif codigo == "50005103" and i < 3:
                    prestador = "ANA CAROLINA OLIVEIRA FARIA FALCAO LTDA"
                elif codigo == "50005103":
                    prestador = "CRIAR E CRESCER TEEN LTDA"
                elif codigo == "50005189" and i == 0:
                    prestador = "PRESTADOR RARO LTDA"
                else:
                    prestador = f"PRESTADOR {codigo} LTDA"
                # O raro (50005189, i==0) fica em CLINICA para permanecer um
                # item individual exato; os demais i==0 exercem a Casa Unimed.
                tipo_prestador = (
                    "ATENDIMENTO ESPECIALIZADO"
                    if i == 0 and codigo not in (COD_AT, "50005189")
                    else "CLINICA ESPECIALIZADA"
                )
                qtd = (
                    2
                    if codigo == "50005103" and mes == "2025/01" and i == 0
                    else 1
                )
                pagamento = (
                    0
                    if codigo == "50005103" and mes == "2025/01" and i == 0
                    else 100 + indice_codigo
                )
                ws.append(
                    [
                        mes,
                        f"MAT-{i + 1}-{codigo}",
                        paciente["nome"],
                        paciente["nasc"],
                        paciente["genero"],
                        f"GUIA-{codigo}-{mes.replace('/', '')}-{i + 1}",
                        date(ano, numero_mes, 15),
                        f"SOL-{i + 1}",
                        "SOLICITANTE RARO" if i == 0 else "SOLICITANTE COMUM",
                        "Psicologo clinico",
                        "SPSADT",
                        int(codigo),
                        f"PROCEDIMENTO {codigo}",
                        "SESSAO",
                        f"PREST-{codigo}-{i + 1}",
                        prestador,
                        tipo_prestador,
                        qtd,
                        pagamento,
                    ]
                )
        wb.save(diretorio / f"fonte-{codigo}.xlsx")


def criar_matriz_sintetica(caminho, registros):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ErroPipeline("openpyxl não instalado") from exc

    esperados = agregados_matriz_esperados(registros)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Resumo_Executivo")
    ws.append(["MATRIZ SINTÉTICA"])
    ws.append([])
    cab = [
        "Terapia", "Código",
        "Registros", "Beneficiários*", "Sessões", "Pagamento",
        "Ben. própria", "Sessões própria", "Média própria",
        "Mediana própria", "Ben. externa", "Sessões externa",
        "Média externa", "Mediana externa", "Custo/sessão própria",
        "Custo/sessão externa",
    ]
    ws.append(cab)
    for terapia_id, valores in sorted(esperados["Resumo_Executivo"].items()):
        definicao = CATALOGO_TERAPIAS[terapia_id]
        ws.append(
            [
                definicao["terapia"],
                " + ".join(definicao["codigos"]),
                *[valor_excel(valores[c]) for c in cab[2:]],
            ]
        )

    ws = wb.create_sheet("Mensal")
    ws.append(["MATRIZ SINTÉTICA"])
    ws.append([])
    cab = [
        "Terapia", "Competência", "Registros", "Beneficiários",
        "Sessões", "Pagamento", "Ben. própria", "Sessões própria",
        "Ben. externa", "Sessões externa",
    ]
    ws.append(cab)
    for (terapia_id, mes), valores in sorted(esperados["Mensal"].items()):
        ws.append(
            [
                CATALOGO_TERAPIAS[terapia_id]["terapia"],
                mes,
                *[valor_excel(valores[c]) for c in cab[2:]],
            ]
        )

    ws = wb.create_sheet("Rede_por_Terapia")
    ws.append(["MATRIZ SINTÉTICA"])
    ws.append([])
    cab = [
        "Terapia", "Rede", "Beneficiários", "Sessões", "Pagamento",
        "Média", "Mediana", "P25", "P75", "Mínimo", "Máximo",
        "Desvio-padrão", "Custo/sessão", "Custo/beneficiário",
        "Sobreposição",
    ]
    ws.append(cab)
    for (terapia_id, rede), valores in sorted(
        esperados["Rede_por_Terapia"].items()
    ):
        ws.append(
            [
                CATALOGO_TERAPIAS[terapia_id]["terapia"],
                rede,
                *[valor_excel(valores[c]) for c in cab[2:]],
            ]
        )
    wb.save(caminho)


def selftest():
    with tempfile.TemporaryDirectory(prefix="tea-f0-selftest-") as tmp:
        raiz = Path(tmp)
        fontes = raiz / "fontes"
        criar_fontes_sinteticas(fontes)
        registros_todos, manifesto = ler_diretorio_fontes(fontes)
        incluidos, mantidas, descartadas = preparar_universo(registros_todos)
        dados, auditoria = construir_dados(
            registros_todos,
            incluidos,
            manifesto,
            mantidas,
            descartadas,
            CORTE_PADRAO,
        )
        matriz = raiz / "Matriz_Analitica_2025_sintetica.xlsx"
        criar_matriz_sintetica(matriz, registros_todos)
        reconciliacao = reconciliar_matriz(matriz, registros_todos)
        dados["meta"]["matriz_2025"] = reconciliacao
        texto = validar_saida(
            dados, auditoria, reconciliacao, exigir_referencias=False
        )

        global_audit = auditoria["recortes"]["tudo"]
        exigir_teste(len(manifesto) == 19, "não leu 19 fontes")
        exigir_teste(len(mantidas) == 17, "não manteve 17 terapias")
        exigir_teste(
            len(descartadas) == 1
            and descartadas[0]["id"] == ID_PECS
            and descartadas[0]["competencias_encontradas"] == 13,
            "cobertura/descarte de PECS incorreto",
        )
        exigir_teste(
            global_audit["metricas"]["n_pacientes"] == 13,
            "deduplicação Nome+Nascimento incorreta ou adulto excluído",
        )
        exigir_teste(
            global_audit["prestadores_observados"]
            == global_audit["prestadores_consolidados"] + 1,
            "merge do par de CNPJ não foi aplicado uma única vez",
        )
        at = dados["meta"]["excecao_at"]
        exigir_teste(
            at["prestador"] == PEDIAKIDS
            and at["terapia_destino"] == "Terapia ABA — Psicologia"
            and at["sessoes"] is not None,
            "exceção AT não ficou rastreável",
        )
        exigir_teste(
            global_audit["metricas"]["sessoes_pagas"]
            < global_audit["metricas"]["sessoes"],
            "sessões com pagamento zero não saíram do denominador",
        )
        denominador_teste = metricas_exatas(
            [
                {"paciente": ("A", "2020-01-01"), "qtd": Decimal("2"),
                 "valor": Decimal("-10")},
                {"paciente": ("B", "2020-01-01"), "qtd": Decimal("3"),
                 "valor": Decimal("0")},
            ]
        )
        exigir_teste(
            denominador_teste["sessoes_pagas"] == Decimal("2"),
            "denominador deve excluir somente pagamento zero",
        )
        exigir_teste(
            len(dados["meta"]["periodos"]) == 30
            and {
                p["tipo"] for p in dados["meta"]["periodos"]
            } == {"tudo", "ano", "semestre", "trimestre", "mes"},
            "períodos permitidos incompletos",
        )
        raro = next(
            x for x in dados["recortes"]["tudo"]["prestadores"]
            if x["nome"] == "PRESTADOR RARO LTDA"
        )
        exigir_teste(
            raro["pacientes"] == 1
            and raro["pagamento"] is not None
            and raro["sessoes"] is not None,
            "prestador raro não permaneceu exato no artefato privado",
        )
        exigir_teste(
            reconciliacao["campos_comparados"] > 0
            and reconciliacao["status"] == "reconciliada",
            "Matriz sintética não reconciliou",
        )

        # Contrato v5: consolidação da Casa Unimed, disciplina × canal e kits.
        recorte_tudo = dados["recortes"]["tudo"]
        resumo_tudo = recorte_tudo["resumo"]
        casa = next(
            (x for x in recorte_tudo["prestadores"] if x["nome"] == CASA_UNIMED),
            None,
        )
        exigir_teste(
            casa is not None and casa.get("composicao"),
            "entidade CASA UNIMED ausente ou sem composição",
        )
        exigir_teste(
            casa["canais"] == [CANAL_CASA_UNIMED]
            and casa["quantidade_entidades"] == len(casa["composicao"]) == 16
            and resumo_tudo["profissionais_casa_unimed"] == 16,
            "composição da Casa Unimed deve consolidar 16 profissionais",
        )
        exigir_teste(
            sum(
                (Decimal(str(p["sessoes"])) for p in casa["composicao"]),
                Decimal(0),
            )
            == Decimal(str(casa["sessoes"])),
            "composição da Casa Unimed não soma as sessões da entidade",
        )
        item_misto = next(
            x for x in recorte_tudo["prestadores"]
            if x["nome"] == NOME_MERGE_CNPJ
        )
        exigir_teste(
            NOME_MERGE_CNPJ in {p["nome"] for p in casa["composicao"]}
            and CANAL_CASA_UNIMED not in item_misto["canais"],
            "caso misto: profissional deve estar na composição da Casa e "
            "permanecer individual sem o canal Casa Unimed",
        )
        exigir_teste(
            CASA_UNIMED in recorte_tudo["serie_prestador"],
            "serie_prestador deve conter CASA UNIMED",
        )
        exigir_teste(
            resumo_tudo["prestadores_publicados"]
            == resumo_tudo["prestadores_consolidados"] + 1
            == len(recorte_tudo["prestadores"]),
            "entidades publicadas devem ser consolidados + 1 na fixture",
        )
        registro_casa = next(
            r for r in incluidos if r["tipo_prest"] == TIPO_CASA_UNIMED
        )
        exigir_teste(
            rede_matriz(registro_casa) == "Rede própria"
            and registro_casa["prest_entidade"] == CASA_UNIMED,
            "rede_matriz e prest_entidade devem coexistir sem derivar um do outro",
        )
        solicitante_raro = next(
            x for x in recorte_tudo["solicitantes"]["itens"]
            if x["nome"] == "SOLICITANTE RARO"
        )
        exigir_teste(
            {e["nome"] for e in solicitante_raro["executores"]}
            == {CASA_UNIMED, PEDIAKIDS, "PRESTADOR RARO LTDA"},
            "executores do solicitante raro devem consolidar a Casa Unimed",
        )
        exigir_teste(
            any(
                fatia["nome"] == CASA_UNIMED
                for fatia in recorte_tudo["prestador_disciplina"]
            ),
            "prestador_disciplina deve conter fatias da entidade CASA UNIMED",
        )
        exigir_teste(
            recorte_tudo["referencias"]["preco_disciplina"][DISC_TO]["n"] == 6,
            "referências devem permanecer na granularidade profissional",
        )
        canal_casa = next(
            x for x in recorte_tudo["canais"] if x["canal"] == CANAL_CASA_UNIMED
        )
        exigir_teste(
            canal_casa["prestadores"] == 1 and canal_casa["profissionais"] == 16,
            "canal Casa Unimed deve contar 1 entidade e 16 profissionais",
        )
        exigir_teste(
            set(recorte_tudo["disciplina_canais"])
            == set(recorte_tudo["resumo_disciplina"])
            == set(recorte_tudo["disciplina_serie"])
            == set(recorte_tudo["pacientes_disciplina"]),
            "disciplinas divergentes entre os agregados por disciplina",
        )
        exigir_teste(
            all(
                {CANAL_CASA_UNIMED, "Rede Credenciada"}
                <= {x["canal"] for x in itens}
                for itens in recorte_tudo["disciplina_canais"].values()
            ),
            "disciplina_canais deve cobrir Casa Unimed e Rede Credenciada",
        )
        exigir_teste(
            recorte_tudo["resumo_disciplina"][DISC_PSI]["pacientes"] == 13
            and recorte_tudo["resumo_disciplina"][DISC_FONO]["pacientes"] == 11,
            "crianças únicas por disciplina divergem da fixture",
        )
        kits = recorte_tudo["kits"]
        exigir_teste(
            set(kits) == {"tripe", "tripe-ampliado"},
            "kits publicados devem ser tripé (k==5) e tripé ampliado (k==6)",
        )
        tripe = kits["tripe"]
        exigir_teste(
            tripe["criancas"] == 5
            and kits["tripe-ampliado"]["criancas"] == 6
            and tripe["crianca_meses"] == 90
            and tripe["sessoes"] == 270,
            "coorte do kit tripé diverge da fixture",
        )
        exigir_teste(
            tripe["custo_crianca_mes"]
            == dinheiro(Decimal(str(tripe["pagamento"])) / Decimal(90))
            and [x["canal"] for x in tripe["canais"]] == ["Rede Credenciada"]
            and tripe["canais"][0]["pagamento"] == tripe["pagamento"],
            "métricas do kit tripé incoerentes",
        )
        valores_guarda = [
            "2025/01", "MAT-GUARDA", "CRIANCA GUARDA", date(2015, 1, 1),
            "FEMININO", "GUIA-GUARDA", date(2025, 1, 15), "SOL-GUARDA",
            "SOLICITANTE GUARDA", "Psicologo clinico", "SPSADT",
            50005103, "PROCEDIMENTO GUARDA", "SESSAO", "PREST-GUARDA",
            "RESINTO ESPACO INTEGRADO LTDA", "ATENDIMENTO ESPECIALIZADO",
            1, 100,
        ]
        try:
            normalizar_registro(valores_guarda, Path("fixture-guarda.xlsx"), 2)
        except ErroPipeline:
            pass
        else:
            raise ErroPipeline(
                "selftest: guarda de coerência da Casa Unimed não bloqueou"
            )
        exigir_teste(
            "CRIANCA SINTETICA" not in texto,
            "nome sintético vazou no JSON",
        )
        exigir_teste(
            primeiro_padrao_encontrado("JOAO DA SILVA", {"JOÃO DA SILVA"}),
            "anti-vazamento não normalizou acentos",
        )
        exigir_teste('"<5"' not in texto, "artefato privado contém máscara <5")
        exigir_teste(
            "OUTROS PRESTADORES" not in texto
            and "OUTRAS TERAPIAS" not in texto
            and "OUTROS SOLICITANTES" not in texto
            and "OUTRAS COMBINACOES" not in texto,
            "artefato privado contém agrupamento OUTROS",
        )
        exigir_teste(
            not valores_equivalentes(
                Decimal("100.01"), Decimal("100.00"), "Pagamento"
            ),
            "reconciliação monetária aceitou diferença de um centavo",
        )
        try:
            validar_anti_vazamento(
                "CRIANCA SINTETICA 1", registros_todos
            )
        except ErroPipeline:
            pass
        else:
            raise ErroPipeline("selftest: mutação anti-vazamento não foi bloqueada")

        falso_git = raiz / "falso-git"
        falso_git.mkdir()
        (falso_git / ".git").mkdir()
        arquivo_interno = falso_git / "fonte.xlsx"
        arquivo_interno.write_bytes(b"fixture")
        try:
            exigir_fora_de_git(arquivo_interno, "fixture")
        except ErroPipeline:
            pass
        else:
            raise ErroPipeline("selftest: insumo dentro de Git não foi bloqueado")

        try:
            exigir_saida_fora_de_git(falso_git / "saida.json")
        except ErroPipeline:
            pass
        else:
            raise ErroPipeline("selftest: saída dentro de Git não foi bloqueada")

        try:
            exigir_saida_fora_de_git(REPO_ROOT / "saida-privada-teste.json")
        except ErroPipeline:
            pass
        else:
            raise ErroPipeline("selftest: saída dentro do repo não foi bloqueada")

        saida_privada = exigir_saida_fora_de_git(raiz / "saida-privada.json")
        escrita_atomica(saida_privada, texto + "\n")
        exigir_teste(
            json.loads(saida_privada.read_text(encoding="utf-8")) == dados,
            "artefato privado gravado diverge do objeto validado",
        )

        # Uma falha na segunda troca deve restaurar a primeira saída.
        saida_a, saida_b = raiz / "saida-a.txt", raiz / "saida-b.txt"
        saida_a.write_text("anterior-a", encoding="utf-8")
        saida_b.write_text("anterior-b", encoding="utf-8")
        escrita_real = globals()["escrita_atomica"]

        def escrita_com_falha(caminho, conteudo):
            if Path(caminho) == saida_b:
                raise OSError("falha sintética na segunda troca")
            return escrita_real(caminho, conteudo)

        globals()["escrita_atomica"] = escrita_com_falha
        try:
            try:
                escrita_atomica_multipla(
                    [(saida_a, "nova-a"), (saida_b, "nova-b")]
                )
            except OSError:
                pass
            else:
                raise ErroPipeline("selftest: falha sintética de escrita não ocorreu")
        finally:
            globals()["escrita_atomica"] = escrita_real
        exigir_teste(
            saida_a.read_text(encoding="utf-8") == "anterior-a"
            and saida_b.read_text(encoding="utf-8") == "anterior-b",
            "rollback não restaurou o par de saídas",
        )

    print(
        "selftest OK — 19 fontes xlsx, schema/aba, cobertura e descarte, "
        "AT, deduplicação, merge, métricas exatas, anti-vazamento, períodos, "
        "destino privado, pagamento zero, Matriz 2025, Casa Unimed "
        "(consolidação por registro, composição, caso misto, guarda), "
        "disciplina × canal, resumo/pacientes por disciplina e kits (k>=5)"
    )


def executar_build(diretorio, matriz, corte, saida):
    destino = exigir_saida_fora_de_git(saida)
    exigir_fora_de_git(matriz, "Matriz Analítica 2025")
    registros_todos, manifesto = ler_diretorio_fontes(diretorio, matriz)
    incluidos, mantidas, descartadas = preparar_universo(registros_todos)
    dados, auditoria = construir_dados(
        registros_todos,
        incluidos,
        manifesto,
        mantidas,
        descartadas,
        corte,
    )
    # A Matriz valida as 18 terapias de origem, inclusive PECS. A publicação
    # continua usando apenas ``incluidos`` (17 terapias com 18/18 competências).
    reconciliacao = reconciliar_matriz(matriz, registros_todos)
    dados["meta"]["matriz_2025"] = reconciliacao
    texto_json = validar_saida(
        dados, auditoria, reconciliacao, exigir_referencias=True
    )
    escrita_atomica(destino, texto_json + "\n")
    imprimir_relatorio(dados, auditoria, reconciliacao, destino)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument(
        "diretorio",
        nargs="?",
        help="diretório externo com exatamente 19 planilhas-fonte xlsx",
    )
    ap.add_argument(
        "--matriz",
        help="Matriz Analítica 2025 externa, obrigatória no build real",
    )
    ap.add_argument(
        "--saida",
        help="arquivo JSON privado, obrigatoriamente fora de qualquer Git",
    )
    ap.add_argument(
        "--corte",
        default=CORTE_PADRAO.isoformat(),
        help="data de corte etário AAAA-MM-DD (padrão: 2026-06-30)",
    )
    ap.add_argument(
        "--emit-p",
        action="store_true",
        help="regenera p/painel-tea/index.html a partir do HTML interno atual",
    )
    ap.add_argument(
        "--selftest",
        action="store_true",
        help="roda a fixture sintética multi-fonte fora do repositório",
    )
    args = ap.parse_args()

    try:
        if args.selftest:
            selftest()
            return
        if args.emit_p and not args.diretorio:
            emitir_p()
            return
        if not args.diretorio:
            ap.error("informe o diretório externo com as 19 fontes")
        if not args.matriz:
            ap.error("--matriz é obrigatória no build real")
        if not args.saida:
            ap.error("--saida é obrigatória no build real e deve ficar fora de Git")
        corte = date.fromisoformat(args.corte)
        executar_build(args.diretorio, args.matriz, corte, args.saida)
        if args.emit_p:
            emitir_p()
    except (ErroPipeline, ValueError, OSError) as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
